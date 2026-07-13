"""
模块：国能 e 招计划追踪测试
用途：验收公告详情地址安全、正文截止时间解析、本机计划表 .xlsx 受控导入，以及受控同步。
对接：chnenergy_client；opportunity_watch_service；fixtures/chnenergy_notice_*.html。
二次开发：禁止真实 HTTP；扩展检索/同步用例时仍须阻断外网并保持字段白名单。
"""

from __future__ import annotations

import io
import json
from datetime import date, datetime, timezone
from pathlib import Path
from urllib.parse import quote

import httpx
import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook
from sqlalchemy import inspect, select
from sqlalchemy.exc import IntegrityError

from pydantic import ValidationError

from app.api.schemas import (
    OpportunityWatchAcceptOut,
    OpportunityWatchHitOut,
    OpportunityWatchPlanImportOut,
    OpportunityWatchPlanOut,
    OpportunityWatchSyncAcceptedOut,
    OpportunityWatchSyncRunOut,
)
from app.core.config import Settings
from app.core.database import SessionLocal, engine
from app.main import app
from app.models.entities import (
    BidOpportunityRow,
    BidSourceHitRow,
    BidSourceSyncRunRow,
    BidWatchPlanRow,
    Workspace,
)
from app.services import opportunity_watch_service

# 任务2 固定错误码字典；非法值必须被 ORM 约束与读模型同时拒绝。
WATCH_ERROR_CODES = (
    "source_unavailable",
    "rate_limited",
    "malformed_response",
    "interrupted",
)

FIXTURES = Path(__file__).resolve().parent / "fixtures"


def test_detail_url_builds_fixed_https_static_path():
    from app.services.chnenergy_client import build_notice_detail_url

    url = build_notice_detail_url(
        infoid="b2363623-ea1e-4cc1-8e2d-0c2d2850b697",
        categorynum="001002001",
        infodate="20260709",
    )
    assert url == (
        "https://www.chnenergybidding.com.cn/bidweb/"
        "001/001002/001002001/20260709/"
        "b2363623-ea1e-4cc1-8e2d-0c2d2850b697.html"
    )


def test_detail_url_from_jump_query_fields():
    from app.services.chnenergy_client import build_notice_detail_url_from_jump

    url = build_notice_detail_url_from_jump(
        "jump.html?infoid=b2363623-ea1e-4cc1-8e2d-0c2d2850b697"
        "&categorynum=001002001&infodate=20260709"
    )
    assert url.endswith(
        "/bidweb/001/001002/001002001/20260709/"
        "b2363623-ea1e-4cc1-8e2d-0c2d2850b697.html"
    )
    assert url.startswith("https://www.chnenergybidding.com.cn/")


def test_detail_url_rejects_external_host():
    from app.services.chnenergy_client import (
        ChnenergyClientError,
        build_notice_detail_url_from_jump,
    )

    with pytest.raises(ChnenergyClientError):
        build_notice_detail_url_from_jump(
            "https://evil.example/jump.html?"
            "infoid=b2363623-ea1e-4cc1-8e2d-0c2d2850b697"
            "&categorynum=001002001&infodate=20260709"
        )


def test_detail_url_rejects_non_uuid_infoid():
    from app.services.chnenergy_client import (
        ChnenergyClientError,
        build_notice_detail_url,
    )

    with pytest.raises(ChnenergyClientError):
        build_notice_detail_url(
            infoid="not-a-uuid",
            categorynum="001002001",
            infodate="20260709",
        )


def test_detail_url_rejects_non_eight_digit_date():
    from app.services.chnenergy_client import (
        ChnenergyClientError,
        build_notice_detail_url,
    )

    with pytest.raises(ChnenergyClientError):
        build_notice_detail_url(
            infoid="b2363623-ea1e-4cc1-8e2d-0c2d2850b697",
            categorynum="001002001",
            infodate="2026-07-09",
        )


def test_detail_url_rejects_non_bid_category():
    from app.services.chnenergy_client import (
        ChnenergyClientError,
        build_notice_detail_url,
    )

    # 中标结果等非 001002 前缀类别
    with pytest.raises(ChnenergyClientError):
        build_notice_detail_url(
            infoid="b2363623-ea1e-4cc1-8e2d-0c2d2850b697",
            categorynum="001006001",
            infodate="20260709",
        )


def test_detail_url_six_digit_category_path_segments():
    """6 位类别按每 3 位追加：/001/001002/日期/id.html，不得重复整段。"""
    from app.services.chnenergy_client import build_notice_detail_url

    url = build_notice_detail_url(
        infoid="b2363623-ea1e-4cc1-8e2d-0c2d2850b697",
        categorynum="001002",
        infodate="20260709",
    )
    assert url == (
        "https://www.chnenergybidding.com.cn/bidweb/"
        "001/001002/20260709/"
        "b2363623-ea1e-4cc1-8e2d-0c2d2850b697.html"
    )


def test_detail_url_twelve_digit_category_path_segments():
    from app.services.chnenergy_client import build_notice_detail_url

    url = build_notice_detail_url(
        infoid="b2363623-ea1e-4cc1-8e2d-0c2d2850b697",
        categorynum="001002001001",
        infodate="20260709",
    )
    assert url == (
        "https://www.chnenergybidding.com.cn/bidweb/"
        "001/001002/001002001/001002001001/20260709/"
        "b2363623-ea1e-4cc1-8e2d-0c2d2850b697.html"
    )


def test_detail_url_rejects_impossible_calendar_date():
    from app.services.chnenergy_client import (
        ChnenergyClientError,
        build_notice_detail_url,
    )

    with pytest.raises(ChnenergyClientError):
        build_notice_detail_url(
            infoid="b2363623-ea1e-4cc1-8e2d-0c2d2850b697",
            categorynum="001002001",
            infodate="20260231",
        )


def test_detail_url_rejects_http_scheme_jump():
    from app.services.chnenergy_client import (
        ChnenergyClientError,
        build_notice_detail_url_from_jump,
    )

    with pytest.raises(ChnenergyClientError):
        build_notice_detail_url_from_jump(
            "http://www.chnenergybidding.com.cn/jump.html?"
            "infoid=b2363623-ea1e-4cc1-8e2d-0c2d2850b697"
            "&categorynum=001002001&infodate=20260709"
        )


def test_extract_invalid_calendar_time_needs_review():
    from app.services.chnenergy_client import extract_notice_times

    html = (
        "<html><body><p>投标文件递交的截止时间（投标截止时间，下同）及开标时间为 "
        "2026-02-30 09:00:00（北京时间）。</p></body></html>"
    )
    result = extract_notice_times(html)
    assert result["extraction_status"] == "needs_review"
    assert result.get("deadline_at_local") in (None, "")


def test_extract_conflicting_deadlines_needs_review():
    from app.services.chnenergy_client import extract_notice_times

    html = (
        "<html><body>"
        "<p>投标截止时间为 2026-07-29 09:00:00。</p>"
        "<p>投标截止时间为 2026-07-30 10:00:00。</p>"
        "</body></html>"
    )
    result = extract_notice_times(html)
    assert result["extraction_status"] == "needs_review"
    assert result.get("deadline_at_local") in (None, "")


def test_extract_separate_deadline_and_opening_clauses():
    """计划 §2.3：独立「投标截止时间为」「开标时间为」条款兼容。"""
    from app.services.chnenergy_client import extract_notice_times

    html = (
        "<html><body><p>"
        "投标截止时间为 2026-07-29 09:00:00；开标时间为 2026-07-29 09:30:00。"
        "</p></body></html>"
    )
    result = extract_notice_times(html)
    assert result["extraction_status"] == "resolved"
    assert result["deadline_at_local"] == "2026-07-29 09:00:00"
    assert result["opening_at_local"] == "2026-07-29 09:30:00"
    assert result["source_timezone"] == "Asia/Shanghai"


def test_extract_deadline_and_opening_from_fixture():
    from app.services.chnenergy_client import extract_notice_times

    html = (FIXTURES / "chnenergy_notice_deadline.html").read_text(encoding="utf-8")
    result = extract_notice_times(html)

    assert result["extraction_status"] == "resolved"
    assert result["deadline_at_local"] == "2026-07-29 09:00:00"
    assert result["opening_at_local"] == "2026-07-29 09:00:00"
    assert result["source_timezone"] == "Asia/Shanghai"
    # 不得采信 script/注释中的假时间
    assert "2099" not in (result["deadline_at_local"] or "")
    assert "2020" not in (result["deadline_at_local"] or "")


def test_extract_needs_review_when_deadline_missing():
    from app.services.chnenergy_client import extract_notice_times

    html = (FIXTURES / "chnenergy_notice_needs_review.html").read_text(
        encoding="utf-8"
    )
    result = extract_notice_times(html)

    assert result["extraction_status"] == "needs_review"
    assert result.get("deadline_at_local") in (None, "")
    assert result["source_timezone"] == "Asia/Shanghai"


def _create_watch_workspace(workspace_id: str) -> None:
    """用途：创建独立工作空间，供计划追踪的归属隔离回归测试使用。"""
    db = SessionLocal()
    try:
        db.add(
            Workspace(
                id=workspace_id,
                name="计划追踪隔离工作空间",
                owner_user_id="user_watch_other",
            )
        )
        db.commit()
    finally:
        db.close()


def _watch_plan(workspace_id: str, plan_id: str, fingerprint: str) -> BidWatchPlanRow:
    """用途：构造最小追踪计划实体，避免测试重复填写与业务无关字段。"""
    return BidWatchPlanRow(
        id=plan_id,
        workspace_id=workspace_id,
        title="某项目招标计划",
        buyer="某招标人",
        scope="建设范围",
        duration="12 个月",
        expected_publish_text="2026 年 7 月",
        remark="测试备注",
        fingerprint=fingerprint,
        enabled=True,
    )


def test_watch_tables_are_created_and_keep_workspace_isolation():
    """用途：验收三张追踪表建表、唯一键与服务读取的工作空间边界。"""
    inspector = inspect(engine)
    table_names = set(inspector.get_table_names())
    assert {"bid_watch_plans", "bid_source_sync_runs", "bid_source_hits"} <= table_names
    assert BidOpportunityRow.__tablename__ == "bid_opportunities"

    other_workspace_id = "ws_watch_other"
    _create_watch_workspace(other_workspace_id)
    db = SessionLocal()
    try:
        local_plan = _watch_plan("ws_local", "watch_plan_local", "fp-local")
        other_plan = _watch_plan(other_workspace_id, "watch_plan_other", "fp-other")
        db.add_all([local_plan, other_plan])
        db.commit()

        assert [item.id for item in opportunity_watch_service.list_watch_plans(db, "ws_local")] == [
            "watch_plan_local"
        ]
        assert opportunity_watch_service.list_watch_plans(db, other_workspace_id)[0].id == "watch_plan_other"

        db.add(_watch_plan("ws_local", "watch_plan_duplicate", "fp-local"))
        with pytest.raises(IntegrityError):
            db.commit()
        db.rollback()
    finally:
        db.close()


def test_lifespan_marks_interrupted_runs_failed_and_keeps_hits():
    """用途：验收应用启动后的未完成运行恢复不删除已写入的命中记录。"""
    db = SessionLocal()
    try:
        plan = _watch_plan("ws_local", "watch_plan_recovery", "fp-recovery")
        queued = BidSourceSyncRunRow(
            id="watch_run_queued",
            workspace_id="ws_local",
            source_name="chnenergy",
            status="queued",
            started_at=datetime(2026, 7, 13, tzinfo=timezone.utc),
        )
        running = BidSourceSyncRunRow(
            id="watch_run_running",
            workspace_id="ws_local",
            source_name="chnenergy",
            status="running",
            started_at=datetime(2026, 7, 13, 1, tzinfo=timezone.utc),
        )
        hit = BidSourceHitRow(
            id="watch_hit_recovery",
            workspace_id="ws_local",
            watch_plan_id=plan.id,
            sync_run_id=queued.id,
            source_name="chnenergy",
            source_info_id="b2363623-ea1e-4cc1-8e2d-0c2d2850b697",
            category_num="001002001",
            source_publish_text="2026-07-09 17:14:11",
            title="某项目招标公告",
            deadline_at_local="2026-07-29 09:00:00",
            opening_at_local="2026-07-29 09:00:00",
            source_timezone="Asia/Shanghai",
            extraction_status="resolved",
        )
        db.add_all([plan, queued, running])
        # 先写入被命中的计划和运行记录，明确外键父记录的落库顺序。
        db.flush()
        db.add(hit)
        db.commit()
    finally:
        db.close()

    # lifespan 必须调用恢复服务，而不是仅让 service 单测通过。
    with TestClient(app) as client:
        assert client.get("/api/health").status_code == 200

    db = SessionLocal()
    try:
        runs = {run.id: run for run in opportunity_watch_service.list_watch_runs(db, "ws_local")}
        assert runs["watch_run_queued"].status == "failed"
        assert runs["watch_run_running"].status == "failed"
        assert runs["watch_run_queued"].error_code == "interrupted"
        assert runs["watch_run_running"].error_code == "interrupted"
        assert runs["watch_run_queued"].finished_at is not None
        assert [item.id for item in opportunity_watch_service.list_watch_hits(db, "ws_local")] == [
            "watch_hit_recovery"
        ]
    finally:
        db.close()


def test_watch_schemas_and_settings_keep_sensitive_fields_outside_contract():
    """用途：冻结追踪读模型和服务端限额，防止 URL/Cookie/原文进入 API 契约。"""
    schema_fields = set()
    for schema in (
        OpportunityWatchPlanOut,
        OpportunityWatchSyncRunOut,
        OpportunityWatchHitOut,
        OpportunityWatchAcceptOut,
    ):
        schema_fields.update(schema.model_fields)
    lowered = {field.lower() for field in schema_fields}
    assert not any(
        forbidden in field
        for field in lowered
        for forbidden in ("cookie", "html", "raw", "response", "url")
    )
    assert {"fingerprint", "error_code", "deadline_at_local", "source_timezone"} <= schema_fields

    settings = Settings()
    # 任务2 固定默认上限：2 MiB、120、5、50、1s、5s、15s、重试 1。
    assert settings.max_opportunity_watch_import_bytes == 2 * 1024 * 1024
    assert settings.max_opportunity_watch_plan_rows == 120
    assert settings.max_opportunity_watch_plans_per_sync == 120
    assert settings.max_opportunity_watch_candidates_per_plan == 5
    assert settings.max_opportunity_watch_detail_pages_per_sync == 50
    assert settings.opportunity_watch_min_interval_seconds == 1
    assert settings.opportunity_watch_connect_timeout_seconds == 5
    assert settings.opportunity_watch_read_timeout_seconds == 15
    assert settings.opportunity_watch_search_retry_count == 1


def test_sync_run_error_code_rejects_unknown_values():
    """用途：验收同步运行 error_code 仅允许固定字典或 NULL，非法值被数据库拒绝。"""
    db = SessionLocal()
    try:
        allowed = BidSourceSyncRunRow(
            id="watch_run_error_ok",
            workspace_id="ws_local",
            source_name="chnenergy",
            status="failed",
            started_at=datetime(2026, 7, 13, 2, tzinfo=timezone.utc),
            finished_at=datetime(2026, 7, 13, 2, 1, tzinfo=timezone.utc),
            error_code="source_unavailable",
        )
        empty = BidSourceSyncRunRow(
            id="watch_run_error_null",
            workspace_id="ws_local",
            source_name="chnenergy",
            status="queued",
            started_at=datetime(2026, 7, 13, 2, 2, tzinfo=timezone.utc),
            error_code=None,
        )
        db.add_all([allowed, empty])
        db.commit()

        for code in WATCH_ERROR_CODES:
            row = BidSourceSyncRunRow(
                id=f"watch_run_error_{code}",
                workspace_id="ws_local",
                source_name="chnenergy",
                status="failed",
                started_at=datetime(2026, 7, 13, 3, tzinfo=timezone.utc),
                finished_at=datetime(2026, 7, 13, 3, 1, tzinfo=timezone.utc),
                error_code=code,
            )
            db.add(row)
            db.commit()

        db.add(
            BidSourceSyncRunRow(
                id="watch_run_error_illegal",
                workspace_id="ws_local",
                source_name="chnenergy",
                status="failed",
                started_at=datetime(2026, 7, 13, 4, tzinfo=timezone.utc),
                finished_at=datetime(2026, 7, 13, 4, 1, tzinfo=timezone.utc),
                error_code="remote_timeout_stacktrace",
            )
        )
        with pytest.raises(IntegrityError):
            db.commit()
        db.rollback()
    finally:
        db.close()


def test_sync_run_out_error_code_schema_is_fixed_dictionary():
    """用途：验收读模型 error_code 收紧为固定字典或 None，拒绝远端原文类字符串。"""
    base = {
        "id": "watch_run_schema",
        "workspace_id": "ws_local",
        "source_name": "chnenergy",
        "status": "failed",
        "started_at": datetime(2026, 7, 13, 5, tzinfo=timezone.utc),
        "finished_at": datetime(2026, 7, 13, 5, 1, tzinfo=timezone.utc),
        "plan_count": 0,
        "candidate_count": 0,
        "detail_page_count": 0,
        "resolved_count": 0,
        "needs_review_count": 0,
        "skipped_count": 0,
        "created_at": datetime(2026, 7, 13, 5, tzinfo=timezone.utc),
        "updated_at": datetime(2026, 7, 13, 5, 1, tzinfo=timezone.utc),
    }
    assert OpportunityWatchSyncRunOut(**base, error_code=None).error_code is None
    for code in WATCH_ERROR_CODES:
        assert OpportunityWatchSyncRunOut(**base, error_code=code).error_code == code
    with pytest.raises(ValidationError):
        OpportunityWatchSyncRunOut(**base, error_code="HTTPError: 500 Internal Server Error")


def test_list_watch_runs_and_hits_keep_workspace_isolation():
    """用途：验收 list_watch_runs / list_watch_hits 严格按 workspace_id 过滤，跨空间不可见。"""
    other_workspace_id = "ws_watch_list_other"
    _create_watch_workspace(other_workspace_id)
    db = SessionLocal()
    try:
        local_plan = _watch_plan("ws_local", "watch_plan_list_local", "fp-list-local")
        other_plan = _watch_plan(other_workspace_id, "watch_plan_list_other", "fp-list-other")
        local_run = BidSourceSyncRunRow(
            id="watch_run_list_local",
            workspace_id="ws_local",
            source_name="chnenergy",
            status="succeeded",
            started_at=datetime(2026, 7, 13, 6, tzinfo=timezone.utc),
            finished_at=datetime(2026, 7, 13, 6, 1, tzinfo=timezone.utc),
            error_code=None,
        )
        other_run = BidSourceSyncRunRow(
            id="watch_run_list_other",
            workspace_id=other_workspace_id,
            source_name="chnenergy",
            status="failed",
            started_at=datetime(2026, 7, 13, 6, 2, tzinfo=timezone.utc),
            finished_at=datetime(2026, 7, 13, 6, 3, tzinfo=timezone.utc),
            error_code="rate_limited",
        )
        db.add_all([local_plan, other_plan, local_run, other_run])
        db.flush()
        local_hit = BidSourceHitRow(
            id="watch_hit_list_local",
            workspace_id="ws_local",
            watch_plan_id=local_plan.id,
            sync_run_id=local_run.id,
            source_name="chnenergy",
            source_info_id="aaaaaaaa-bbbb-cccc-dddd-eeeeeeeeeeee",
            category_num="001002001",
            source_publish_text="2026-07-09 10:00:00",
            title="本空间命中",
            deadline_at_local="2026-07-29 09:00:00",
            opening_at_local="2026-07-29 09:00:00",
            source_timezone="Asia/Shanghai",
            extraction_status="resolved",
        )
        other_hit = BidSourceHitRow(
            id="watch_hit_list_other",
            workspace_id=other_workspace_id,
            watch_plan_id=other_plan.id,
            sync_run_id=other_run.id,
            source_name="chnenergy",
            source_info_id="ffffffff-1111-2222-3333-444444444444",
            category_num="001002001",
            source_publish_text="2026-07-09 11:00:00",
            title="他空间命中",
            deadline_at_local=None,
            opening_at_local=None,
            source_timezone="Asia/Shanghai",
            extraction_status="needs_review",
        )
        db.add_all([local_hit, other_hit])
        db.commit()

        assert [item.id for item in opportunity_watch_service.list_watch_runs(db, "ws_local")] == [
            "watch_run_list_local"
        ]
        assert [
            item.id for item in opportunity_watch_service.list_watch_runs(db, other_workspace_id)
        ] == ["watch_run_list_other"]
        assert [item.id for item in opportunity_watch_service.list_watch_hits(db, "ws_local")] == [
            "watch_hit_list_local"
        ]
        assert [
            item.id for item in opportunity_watch_service.list_watch_hits(db, other_workspace_id)
        ] == ["watch_hit_list_other"]
        # 空/未知工作空间必须返回空列表，不得泄漏其他空间数据。
        assert opportunity_watch_service.list_watch_runs(db, "ws_missing") == []
        assert opportunity_watch_service.list_watch_hits(db, "ws_missing") == []
    finally:
        db.close()


def _make_hit(
    *,
    hit_id: str,
    workspace_id: str,
    watch_plan_id: str,
    sync_run_id: str,
    source_info_id: str,
) -> BidSourceHitRow:
    """用途：构造最小命中实体，便于跨空间外键拒绝回归。"""
    return BidSourceHitRow(
        id=hit_id,
        workspace_id=workspace_id,
        watch_plan_id=watch_plan_id,
        sync_run_id=sync_run_id,
        source_name="chnenergy",
        source_info_id=source_info_id,
        category_num="001002001",
        source_publish_text="2026-07-09 12:00:00",
        title="跨空间关系约束样本",
        deadline_at_local="2026-07-29 09:00:00",
        opening_at_local="2026-07-29 09:00:00",
        source_timezone="Asia/Shanghai",
        extraction_status="resolved",
    )


def test_hit_rejects_cross_workspace_plan_or_run_reference():
    """
    用途：验收命中、计划、运行必须同属一个 workspace；跨空间计划或运行引用被数据库拒绝。
    对接：BidSourceHitRow 复合外键；(workspace_id, watch_plan_id)/(workspace_id, sync_run_id)。
    二次开发：不得仅靠服务层过滤；非法组合必须在 ORM/SQLite 层失败。
    """
    other_workspace_id = "ws_watch_fk_other"
    _create_watch_workspace(other_workspace_id)
    db = SessionLocal()
    try:
        local_plan = _watch_plan("ws_local", "watch_plan_fk_local", "fp-fk-local")
        other_plan = _watch_plan(other_workspace_id, "watch_plan_fk_other", "fp-fk-other")
        local_run = BidSourceSyncRunRow(
            id="watch_run_fk_local",
            workspace_id="ws_local",
            source_name="chnenergy",
            status="succeeded",
            started_at=datetime(2026, 7, 13, 7, tzinfo=timezone.utc),
            finished_at=datetime(2026, 7, 13, 7, 1, tzinfo=timezone.utc),
        )
        other_run = BidSourceSyncRunRow(
            id="watch_run_fk_other",
            workspace_id=other_workspace_id,
            source_name="chnenergy",
            status="succeeded",
            started_at=datetime(2026, 7, 13, 7, 2, tzinfo=timezone.utc),
            finished_at=datetime(2026, 7, 13, 7, 3, tzinfo=timezone.utc),
        )
        db.add_all([local_plan, other_plan, local_run, other_run])
        db.flush()

        # 合法：命中、计划、运行均属 ws_local。
        db.add(
            _make_hit(
                hit_id="watch_hit_fk_ok",
                workspace_id="ws_local",
                watch_plan_id=local_plan.id,
                sync_run_id=local_run.id,
                source_info_id="11111111-2222-3333-4444-555555555555",
            )
        )
        db.commit()

        # 计划跨空间：hit.workspace=ws_local，但 watch_plan 属于 other。
        db.add(
            _make_hit(
                hit_id="watch_hit_fk_bad_plan",
                workspace_id="ws_local",
                watch_plan_id=other_plan.id,
                sync_run_id=local_run.id,
                source_info_id="aaaaaaaa-bbbb-cccc-dddd-000000000001",
            )
        )
        with pytest.raises(IntegrityError):
            db.commit()
        db.rollback()

        # 运行跨空间：hit.workspace=ws_local，但 sync_run 属于 other。
        db.add(
            _make_hit(
                hit_id="watch_hit_fk_bad_run",
                workspace_id="ws_local",
                watch_plan_id=local_plan.id,
                sync_run_id=other_run.id,
                source_info_id="aaaaaaaa-bbbb-cccc-dddd-000000000002",
            )
        )
        with pytest.raises(IntegrityError):
            db.commit()
        db.rollback()
    finally:
        db.close()


# ---------- P9B 任务3：计划表 .xlsx 受控导入 ----------

_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_PLAN_HEADERS = (
    "招标计划名称",
    "招标人",
    "范围",
    "计划工期",
    "预计发布公告时间",
    "备注",
)


def _workbook_bytes(
    rows: list[list[object]],
    *,
    header_row_index: int = 3,
    headers: tuple[str, ...] | list[str] = _PLAN_HEADERS,
) -> bytes:
    """用途：在内存构造国能计划表样式工作簿，供导入契约测试使用，不落盘。"""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    # 前两行说明文字，表头默认落在第 3 行，与日常计划表一致。
    for idx in range(1, header_row_index):
        ws.cell(idx, 1, f"说明行 {idx}")
    for col, name in enumerate(headers, start=1):
        ws.cell(header_row_index, col, name)
    for offset, row in enumerate(rows):
        for col, value in enumerate(row, start=1):
            ws.cell(header_row_index + 1 + offset, col, value)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _post_plan_import(
    client: TestClient,
    content: bytes,
    *,
    filename: str = "plans.xlsx",
    workspace_id: str | None = None,
):
    """用途：统一调用计划导入路由，可选切换工作空间请求头。"""
    headers = {}
    if workspace_id is not None:
        headers["X-Workspace-Id"] = workspace_id
    return client.post(
        "/api/opportunity-watch/plans/import",
        files={"file": (filename, content, _XLSX_MIME)},
        headers=headers,
    )


def _count_watch_plans(workspace_id: str) -> int:
    """用途：统计指定工作空间已写入的追踪计划数量。"""
    db = SessionLocal()
    try:
        return len(
            list(
                db.scalars(
                    select(BidWatchPlanRow).where(BidWatchPlanRow.workspace_id == workspace_id)
                ).all()
            )
        )
    finally:
        db.close()


def test_plan_import_out_schema_only_exposes_counts():
    """用途：冻结导入成功响应只含 inserted/skipped/total，禁止文件与远端字段。"""
    fields = set(OpportunityWatchPlanImportOut.model_fields)
    assert fields == {"inserted", "skipped", "total"}
    lowered = {name.lower() for name in fields}
    assert not any(
        forbidden in name
        for name in lowered
        for forbidden in ("cookie", "url", "path", "file", "html", "raw")
    )
    assert OpportunityWatchPlanImportOut(inserted=2, skipped=0, total=2).model_dump() == {
        "inserted": 2,
        "skipped": 0,
        "total": 2,
    }


def test_plan_import_service_inserts_and_skips_duplicates_on_reimport():
    """用途：验收表头在前十行、有效计划写入，以及重复导入按指纹跳过。"""
    content = _workbook_bytes(
        [
            ["计划甲", "招标人甲", "范围甲", "12个月", "2026年7月", "备注甲"],
            ["计划乙", "招标人乙", "范围乙", "6个月", "2026年8月", "备注乙"],
        ]
    )
    db = SessionLocal()
    try:
        first = opportunity_watch_service.import_watch_plans_from_xlsx(
            db,
            "ws_local",
            filename="plans.xlsx",
            content=content,
            max_rows=120,
        )
        assert first == {"inserted": 2, "skipped": 0, "total": 2}
        second = opportunity_watch_service.import_watch_plans_from_xlsx(
            db,
            "ws_local",
            filename="plans.xlsx",
            content=content,
            max_rows=120,
        )
        assert second == {"inserted": 0, "skipped": 2, "total": 2}
        plans = opportunity_watch_service.list_watch_plans(db, "ws_local")
        assert len(plans) == 2
        titles = {item.title for item in plans}
        assert titles == {"计划甲", "计划乙"}
        assert all(item.fingerprint for item in plans)
    finally:
        db.close()


def test_plan_import_service_rejects_missing_title_header_with_zero_write():
    """用途：缺少招标计划名称表头时整批零写入。"""
    content = _workbook_bytes(
        [["计划甲", "招标人甲", "范围甲", "12个月", "2026年7月", "备注"]],
        headers=("计划名称", "招标人", "范围", "计划工期", "预计发布公告时间", "备注"),
    )
    db = SessionLocal()
    try:
        with pytest.raises(opportunity_watch_service.WatchPlanImportValidationError) as exc_info:
            opportunity_watch_service.import_watch_plans_from_xlsx(
                db,
                "ws_local",
                filename="plans.xlsx",
                content=content,
                max_rows=120,
            )
        assert exc_info.value.errors
        assert any("招标计划名称" in str(item.get("message", "")) for item in exc_info.value.errors)
        assert _count_watch_plans("ws_local") == 0
    finally:
        db.close()


def test_plan_import_service_rejects_blank_title_on_partial_row():
    """用途：非空行缺计划名返回 Excel 实际行号，整批零写入；全空白行仍跳过。"""
    content = _workbook_bytes(
        [
            ["", "招标人甲", "范围甲", "", "", ""],
            [None, None, None, None, None, None],
            ["计划乙", "招标人乙", "范围乙", "6个月", "2026年8月", "备注乙"],
        ]
    )
    db = SessionLocal()
    try:
        with pytest.raises(opportunity_watch_service.WatchPlanImportValidationError) as exc_info:
            opportunity_watch_service.import_watch_plans_from_xlsx(
                db,
                "ws_local",
                filename="plans.xlsx",
                content=content,
                max_rows=120,
            )
        # 表头在第 3 行，首条数据为 Excel 第 4 行。
        assert any(
            item.get("row") == 4 and item.get("field") == "招标计划名称"
            for item in exc_info.value.errors
        )
        assert _count_watch_plans("ws_local") == 0
    finally:
        db.close()


def test_plan_import_service_rejects_when_plan_rows_exceed_limit():
    """用途：超过 max_rows 计划行时拒绝写入。"""
    rows = [[f"计划{i}", "招标人", "范围", "1个月", "2026年7月", ""] for i in range(121)]
    content = _workbook_bytes(rows)
    db = SessionLocal()
    try:
        with pytest.raises(ValueError, match="120"):
            opportunity_watch_service.import_watch_plans_from_xlsx(
                db,
                "ws_local",
                filename="plans.xlsx",
                content=content,
                max_rows=120,
            )
        assert _count_watch_plans("ws_local") == 0
    finally:
        db.close()


def test_plan_import_service_stops_iterating_after_max_rows(monkeypatch: pytest.MonkeyPatch):
    """用途：超过 max_rows 时立即停止枚举行，禁止先物化全表再截断。"""
    max_rows = 3
    total_data_rows = 80
    yielded = {"count": 0}

    header = list(_PLAN_HEADERS)
    data_rows = [
        [f"计划{i}", "招标人", "范围", "1个月", "2026年7月", ""] for i in range(total_data_rows)
    ]

    class _CountingSheet:
        """可控假工作表：记录 iter_rows 实际产出次数。"""

        def iter_rows(self, *args, **kwargs):
            yielded["count"] += 1
            yield header
            for row in data_rows:
                yielded["count"] += 1
                yield row

    class _CountingWorkbook:
        worksheets = [_CountingSheet()]

        def close(self) -> None:
            return None

    monkeypatch.setattr(
        opportunity_watch_service,
        "load_workbook",
        lambda *args, **kwargs: _CountingWorkbook(),
    )

    db = SessionLocal()
    try:
        with pytest.raises(ValueError, match=str(max_rows)):
            opportunity_watch_service.import_watch_plans_from_xlsx(
                db,
                "ws_local",
                filename="plans.xlsx",
                content=b"fake-xlsx-bytes",
                max_rows=max_rows,
            )
        # 表头 1 行 + 上限内计划行 + 触发上限的第 max_rows+1 行，不得继续枚举剩余行。
        assert yielded["count"] == 1 + max_rows + 1
        assert yielded["count"] < total_data_rows
        assert _count_watch_plans("ws_local") == 0
    finally:
        db.close()


def test_plan_import_service_skips_duplicate_fingerprints_in_same_batch():
    """用途：同批重复计划名+招标人+范围只计 skipped，不得产生重复行。"""
    content = _workbook_bytes(
        [
            ["计划甲", "招标人甲", "范围甲", "12个月", "2026年7月", "备注1"],
            ["计划甲", "招标人甲", "范围甲", "24个月", "2026年8月", "备注2"],
        ]
    )
    db = SessionLocal()
    try:
        result = opportunity_watch_service.import_watch_plans_from_xlsx(
            db,
            "ws_local",
            filename="plans.xlsx",
            content=content,
            max_rows=120,
        )
        assert result == {"inserted": 1, "skipped": 1, "total": 2}
        plans = opportunity_watch_service.list_watch_plans(db, "ws_local")
        assert len(plans) == 1
        assert plans[0].title == "计划甲"
        assert plans[0].duration == "12个月"
    finally:
        db.close()


def test_plan_import_service_allows_same_plan_across_workspaces():
    """用途：跨工作空间允许各自导入同一计划指纹。"""
    other_workspace_id = "ws_watch_import_other"
    _create_watch_workspace(other_workspace_id)
    content = _workbook_bytes(
        [["共享计划", "共享招标人", "共享范围", "12个月", "2026年7月", "备注"]]
    )
    db = SessionLocal()
    try:
        local = opportunity_watch_service.import_watch_plans_from_xlsx(
            db,
            "ws_local",
            filename="plans.xlsx",
            content=content,
            max_rows=120,
        )
        other = opportunity_watch_service.import_watch_plans_from_xlsx(
            db,
            other_workspace_id,
            filename="plans.xlsx",
            content=content,
            max_rows=120,
        )
        assert local == {"inserted": 1, "skipped": 0, "total": 1}
        assert other == {"inserted": 1, "skipped": 0, "total": 1}
        assert _count_watch_plans("ws_local") == 1
        assert _count_watch_plans(other_workspace_id) == 1
    finally:
        db.close()


def test_plan_import_api_success_and_reimport(client: TestClient):
    """用途：验收导入路由 201 响应契约与重复上传跳过。"""
    content = _workbook_bytes(
        [
            ["计划甲", "招标人甲", "范围甲", "12个月", "2026年7月", "备注甲"],
            ["计划乙", "招标人乙", "范围乙", "6个月", "2026年8月", "备注乙"],
        ]
    )
    first = _post_plan_import(client, content)
    assert first.status_code == 201
    assert first.json() == {"inserted": 2, "skipped": 0, "total": 2}
    second = _post_plan_import(client, content)
    assert second.status_code == 201
    assert second.json() == {"inserted": 0, "skipped": 2, "total": 2}
    assert _count_watch_plans("ws_local") == 2


def test_plan_import_api_isolates_workspaces(client: TestClient):
    """用途：验收跨工作空间导入互不影响。"""
    other_workspace_id = "ws_watch_api_other"
    _create_watch_workspace(other_workspace_id)
    content = _workbook_bytes(
        [["跨空间计划", "招标人", "范围", "12个月", "2026年7月", ""]]
    )
    local = _post_plan_import(client, content)
    other = _post_plan_import(client, content, workspace_id=other_workspace_id)
    assert local.status_code == 201
    assert other.status_code == 201
    assert local.json() == {"inserted": 1, "skipped": 0, "total": 1}
    assert other.json() == {"inserted": 1, "skipped": 0, "total": 1}
    assert _count_watch_plans("ws_local") == 1
    assert _count_watch_plans(other_workspace_id) == 1


def test_plan_import_api_validation_errors_zero_write(client: TestClient):
    """用途：缺表头与错误数据行返回 422，当前工作空间零写入。"""
    missing_header = _workbook_bytes(
        [["计划甲", "招标人", "范围", "1个月", "2026年7月", ""]],
        headers=("计划名称", "招标人", "范围", "计划工期", "预计发布公告时间", "备注"),
    )
    resp = _post_plan_import(client, missing_header)
    assert resp.status_code == 422
    body = resp.json()
    assert "detail" in body
    assert _count_watch_plans("ws_local") == 0

    blank_title = _workbook_bytes(
        [["", "招标人甲", "范围甲", "1个月", "2026年7月", "备注"]]
    )
    resp2 = _post_plan_import(client, blank_title)
    assert resp2.status_code == 422
    detail = resp2.json()["detail"]
    errors = detail["errors"] if isinstance(detail, dict) else detail
    assert any(
        (item.get("row") == 4 and item.get("field") == "招标计划名称")
        for item in errors
    )
    assert _count_watch_plans("ws_local") == 0


def test_plan_import_api_rejects_non_xlsx_and_oversized(client: TestClient):
    """用途：非 .xlsx 扩展名与超 2MiB 文件返回 400，且不写入计划。"""
    valid_content = _workbook_bytes(
        [["计划甲", "招标人", "范围", "1个月", "2026年7月", ""]]
    )
    for filename in ("plans.csv", "plans.json", "plans.xlsx.exe", "plans"):
        resp = _post_plan_import(client, valid_content, filename=filename)
        assert resp.status_code == 400, filename
    assert _count_watch_plans("ws_local") == 0

    oversized = b"A" * (2 * 1024 * 1024 + 1)
    resp = _post_plan_import(client, oversized, filename="plans.xlsx")
    assert resp.status_code == 400
    assert _count_watch_plans("ws_local") == 0


# ---------- P9B 任务4：国能 e 招受控同步 ----------

_PORTAL_URL = "https://www.chnenergybidding.com.cn/bidweb/"
_SEARCH_URL = (
    "https://www.chnenergybidding.com.cn/bidfulltextsearch/rest/"
    "inteligentSearch/getFullTextData"
)
_INFO_ID_A = "b2363623-ea1e-4cc1-8e2d-0c2d2850b697"
_INFO_ID_B = "c3474734-fb2f-5dd2-9f3e-1d3e3961c7a8"
_INFO_ID_C = "d4585845-0c30-6ee3-a04f-2e4f4a72d8b9"


def _jump(
    infoid: str,
    categorynum: str = "001002001",
    infodate: str = "20260709",
) -> str:
    return f"jump.html?infoid={infoid}&categorynum={categorynum}&infodate={infodate}"


def _deadline_html() -> str:
    return (FIXTURES / "chnenergy_notice_deadline.html").read_text(encoding="utf-8")


def _needs_review_html() -> str:
    return (FIXTURES / "chnenergy_notice_needs_review.html").read_text(encoding="utf-8")


def _seed_enabled_plan(
    *,
    plan_id: str = "watch_plan_sync_a",
    title: str = "某项目招标计划",
    fingerprint: str = "fp-sync-a",
    workspace_id: str = "ws_local",
    enabled: bool = True,
) -> None:
    db = SessionLocal()
    try:
        db.add(
            BidWatchPlanRow(
                id=plan_id,
                workspace_id=workspace_id,
                title=title,
                buyer="某招标人",
                scope="建设范围",
                duration="12 个月",
                expected_publish_text="2026 年 7 月",
                remark="同步测试",
                fingerprint=fingerprint,
                enabled=enabled,
            )
        )
        db.commit()
    finally:
        db.close()


def _mock_transport(handler):
    """用途：包装 MockTransport；handler 内未声明请求应 AssertionError，阻断外网。"""
    return httpx.MockTransport(handler)


def test_build_fixed_search_body_matches_verified_chnenergy_template():
    """用途：整包比对已核验的国能 e 招检索 JSON 冻结模板（仅 wd 随计划名变化）。"""
    from app.services.chnenergy_client import build_fixed_search_body

    plan_title = "某项目招标计划"
    expected = {
        "token": "",
        "pn": 0,
        "rn": 5,
        "sdt": "",
        "edt": "",
        "wd": quote(plan_title, safe=""),
        "inc_wd": "",
        "exc_wd": "",
        "fields": "title;content",
        "cnum": "",
        "sort": '{"infodate":0}',
        "ssort": "title",
        "cl": 500,
        "terminal": "",
        "condition": None,
        "time": None,
        "highlights": "title;content",
        "statistics": None,
        "unionCondition": None,
        "accuracy": "",
        "noParticiple": "1",
        "searchRange": None,
    }
    body = build_fixed_search_body(plan_title)
    assert body == expected
    assert "isBusiness" not in body
    assert set(body.keys()) == set(expected.keys())


def test_sync_client_portal_search_contract_and_limit():
    """用途：门户 uid + 检索 HTTPS/固定 Referer/JSON/rn=5/最多五候选。"""
    from app.services import chnenergy_client as client_mod

    calls: list[httpx.Request] = []

    def handler(request: httpx.Request) -> httpx.Response:
        calls.append(request)
        url = str(request.url)
        if request.method == "GET" and url.rstrip("/") == _PORTAL_URL.rstrip("/"):
            return httpx.Response(
                200,
                headers=[("set-cookie", "uid=mock-uid-token; Path=/")],
                text="<html>portal</html>",
            )
        if request.method == "POST" and url == _SEARCH_URL:
            body = json.loads(request.content.decode("utf-8"))
            assert body["rn"] == 5
            assert body["wd"] == quote("某项目招标计划", safe="")
            assert body["fields"] == "title;content"
            # 整包字段与已核验模板一致，禁止 isBusiness 等漂移字段
            expected_body = client_mod.build_fixed_search_body("某项目招标计划")
            assert body == expected_body
            assert "isBusiness" not in body
            assert body["sort"] == '{"infodate":0}'
            assert body["condition"] is None
            assert body["time"] is None
            assert body["highlights"] == "title;content"
            assert body["noParticiple"] == "1"
            assert request.headers.get("referer") == _PORTAL_URL
            assert request.url.scheme == "https"
            records = [
                {
                    "title": f"候选{i}",
                    "infodate": "2026-07-09 10:00:00",
                    "linkurl": _jump(f"b2363623-ea1e-4cc1-8e2d-{i:012d}"),
                }
                for i in range(7)
            ]
            return httpx.Response(200, json={"result": {"records": records}})
        raise AssertionError(f"未声明请求: {request.method} {url}")

    sleeps: list[float] = []
    controlled = client_mod.ChnenergyControlledClient(
        transport=_mock_transport(handler),
        sleep_fn=lambda seconds: sleeps.append(seconds),
        min_interval_seconds=1.0,
    )
    with controlled:
        controlled.ensure_session()
        records = controlled.search_candidates("某项目招标计划")
    assert len(records) == 5
    assert all(set(item) <= {"title", "infodate", "linkurl"} for item in records)
    assert calls[0].method == "GET"
    assert calls[1].method == "POST"
    assert controlled.follow_redirects is False
    assert sleeps  # 门户与检索之间触发限频


def test_sync_client_missing_uid_is_source_unavailable():
    from app.services import chnenergy_client as client_mod

    def handler(request: httpx.Request) -> httpx.Response:
        if request.method == "GET":
            return httpx.Response(200, text="<html>no cookie</html>")
        raise AssertionError("无 uid 后不得继续请求")

    controlled = client_mod.ChnenergyControlledClient(
        transport=_mock_transport(handler),
        sleep_fn=lambda _s: None,
        min_interval_seconds=0,
    )
    with controlled:
        with pytest.raises(client_mod.ChnenergySyncStopError) as exc:
            controlled.ensure_session()
        assert exc.value.error_code == "source_unavailable"


def test_sync_client_403_and_429_are_rate_limited():
    from app.services import chnenergy_client as client_mod

    for status_code in (403, 429):

        def handler(request: httpx.Request, code: int = status_code) -> httpx.Response:
            if request.method == "GET":
                return httpx.Response(
                    200,
                    headers=[("set-cookie", "uid=abc; Path=/")],
                    text="ok",
                )
            return httpx.Response(code, text="denied")

        controlled = client_mod.ChnenergyControlledClient(
            transport=_mock_transport(handler),
            sleep_fn=lambda _s: None,
            min_interval_seconds=0,
        )
        with controlled:
            controlled.ensure_session()
            with pytest.raises(client_mod.ChnenergySyncStopError) as exc:
                controlled.search_candidates("计划")
            assert exc.value.error_code == "rate_limited"


def test_sync_client_malformed_search_payload():
    from app.services import chnenergy_client as client_mod

    def handler(request: httpx.Request) -> httpx.Response:
        if request.method == "GET":
            return httpx.Response(
                200,
                headers=[("set-cookie", "uid=abc; Path=/")],
                text="ok",
            )
        return httpx.Response(200, json={"result": {"items": []}})

    controlled = client_mod.ChnenergyControlledClient(
        transport=_mock_transport(handler),
        sleep_fn=lambda _s: None,
        min_interval_seconds=0,
    )
    with controlled:
        controlled.ensure_session()
        with pytest.raises(client_mod.ChnenergySyncStopError) as exc:
            controlled.search_candidates("计划")
        assert exc.value.error_code == "malformed_response"


def test_sync_create_run_rejects_concurrent_and_isolates_workspace():
    """用途：同空间 queued/running 冲突；跨空间不可查询运行。"""
    _seed_enabled_plan()
    db = SessionLocal()
    try:
        first = opportunity_watch_service.create_watch_sync_run(db, "ws_local")
        assert first.status == "queued"
        assert first.source_name == "chnenergy"
        assert first.plan_count == 1
        with pytest.raises(opportunity_watch_service.WatchSyncConflictError):
            opportunity_watch_service.create_watch_sync_run(db, "ws_local")
    finally:
        db.close()

    other = "ws_sync_other"
    _create_watch_workspace(other)
    _seed_enabled_plan(
        plan_id="watch_plan_sync_other",
        fingerprint="fp-sync-other",
        workspace_id=other,
        title="他空间计划",
    )
    db = SessionLocal()
    try:
        other_run = opportunity_watch_service.create_watch_sync_run(db, other)
        assert other_run.workspace_id == other
        assert (
            opportunity_watch_service.get_watch_sync_run(db, "ws_local", other_run.id)
            is None
        )
        assert (
            opportunity_watch_service.get_watch_sync_run(db, other, other_run.id)
            is not None
        )
    finally:
        db.close()


def test_sync_execute_happy_path_resolved_and_filters_non_bid():
    """用途：resolved 命中写入；非 001002/非法 linkurl 跳过；不自动立项。"""
    _seed_enabled_plan(title="同步计划甲")
    detail_hits = {"count": 0}

    def handler(request: httpx.Request) -> httpx.Response:
        url = str(request.url)
        if request.method == "GET" and url.rstrip("/") == _PORTAL_URL.rstrip("/"):
            return httpx.Response(
                200,
                headers=[("set-cookie", "uid=ok; Path=/")],
                text="portal",
            )
        if request.method == "POST" and url == _SEARCH_URL:
            body = json.loads(request.content.decode("utf-8"))
            assert body["wd"] == quote("同步计划甲", safe="")
            return httpx.Response(
                200,
                json={
                    "result": {
                        "records": [
                            {
                                "title": "可解析招标公告",
                                "infodate": "2026-07-09 17:14:11",
                                "linkurl": _jump(_INFO_ID_A, "001002001", "20260709"),
                            },
                            {
                                "title": "中标结果应跳过",
                                "infodate": "2026-07-09 18:00:00",
                                "linkurl": _jump(_INFO_ID_B, "001006001", "20260709"),
                            },
                            {
                                "title": "非法跳转",
                                "infodate": "2026-07-09 19:00:00",
                                "linkurl": "https://evil.example/x",
                            },
                        ]
                    }
                },
            )
        if request.method == "GET" and url.endswith(f"{_INFO_ID_A}.html"):
            detail_hits["count"] += 1
            assert url.startswith("https://www.chnenergybidding.com.cn/bidweb/")
            assert "evil" not in url
            return httpx.Response(200, text=_deadline_html())
        if _INFO_ID_B in url:
            raise AssertionError("非招标类别不得读详情")
        raise AssertionError(f"未声明请求: {request.method} {url}")

    db = SessionLocal()
    try:
        run_id = opportunity_watch_service.create_watch_sync_run(db, "ws_local").id
    finally:
        db.close()

    opportunity_watch_service.execute_sync_run(
        run_id,
        transport=_mock_transport(handler),
        sleep_fn=lambda _s: None,
    )

    db = SessionLocal()
    try:
        run = opportunity_watch_service.get_watch_sync_run(db, "ws_local", run_id)
        assert run is not None
        assert run.status == "succeeded"
        assert run.error_code is None
        assert run.plan_count == 1
        assert run.candidate_count == 3
        assert run.detail_page_count == 1
        assert run.resolved_count == 1
        assert run.needs_review_count == 0
        assert run.skipped_count == 2
        assert detail_hits["count"] == 1
        hits = opportunity_watch_service.list_watch_hits(db, "ws_local")
        assert len(hits) == 1
        hit = hits[0]
        assert hit.source_info_id == _INFO_ID_A
        assert hit.category_num == "001002001"
        assert hit.deadline_at_local == "2026-07-29 09:00:00"
        assert hit.opening_at_local == "2026-07-29 09:00:00"
        assert hit.extraction_status == "resolved"
        assert hit.source_timezone == "Asia/Shanghai"
        assert hit.accepted_opportunity_id is None
        assert list(db.scalars(select(BidOpportunityRow)).all()) == []
        dumped = json.dumps(
            OpportunityWatchSyncRunOut.model_validate(run).model_dump(
                by_alias=True, mode="json"
            )
        ).lower()
        assert "cookie" not in dumped
        assert "http" not in dumped
    finally:
        db.close()


def test_sync_execute_needs_review_on_detail_without_deadline():
    _seed_enabled_plan(title="待复核计划")

    def handler(request: httpx.Request) -> httpx.Response:
        url = str(request.url)
        if request.method == "GET" and url.rstrip("/") == _PORTAL_URL.rstrip("/"):
            return httpx.Response(
                200,
                headers=[("set-cookie", "uid=ok; Path=/")],
                text="portal",
            )
        if request.method == "POST":
            return httpx.Response(
                200,
                json={
                    "result": {
                        "records": [
                            {
                                "title": "无时间公告",
                                "infodate": "2026-07-09 17:14:11",
                                "linkurl": _jump(_INFO_ID_A),
                            }
                        ]
                    }
                },
            )
        if request.method == "GET" and url.endswith(".html"):
            return httpx.Response(200, text=_needs_review_html())
        raise AssertionError(url)

    db = SessionLocal()
    try:
        run_id = opportunity_watch_service.create_watch_sync_run(db, "ws_local").id
    finally:
        db.close()

    opportunity_watch_service.execute_sync_run(
        run_id,
        transport=_mock_transport(handler),
        sleep_fn=lambda _s: None,
    )
    db = SessionLocal()
    try:
        run = opportunity_watch_service.get_watch_sync_run(db, "ws_local", run_id)
        assert run is not None
        assert run.status == "succeeded"
        assert run.needs_review_count == 1
        assert run.resolved_count == 0
        hit = opportunity_watch_service.list_watch_hits(db, "ws_local")[0]
        assert hit.extraction_status == "needs_review"
        assert hit.deadline_at_local in (None, "")
    finally:
        db.close()


def test_sync_execute_reuses_infoid_detail_within_run():
    """用途：同运行同 infoid 只读一次详情；两计划各保留一条命中。"""
    _seed_enabled_plan(plan_id="watch_plan_sync_1", title="计划一", fingerprint="fp1")
    _seed_enabled_plan(plan_id="watch_plan_sync_2", title="计划二", fingerprint="fp2")
    detail_hits = {"count": 0}

    def handler(request: httpx.Request) -> httpx.Response:
        url = str(request.url)
        if request.method == "GET" and url.rstrip("/") == _PORTAL_URL.rstrip("/"):
            return httpx.Response(
                200,
                headers=[("set-cookie", "uid=ok; Path=/")],
                text="portal",
            )
        if request.method == "POST" and url == _SEARCH_URL:
            return httpx.Response(
                200,
                json={
                    "result": {
                        "records": [
                            {
                                "title": "共享公告",
                                "infodate": "2026-07-09 17:14:11",
                                "linkurl": _jump(_INFO_ID_A),
                            }
                        ]
                    }
                },
            )
        if request.method == "GET" and url.endswith(f"{_INFO_ID_A}.html"):
            detail_hits["count"] += 1
            return httpx.Response(200, text=_deadline_html())
        raise AssertionError(f"{request.method} {url}")

    db = SessionLocal()
    try:
        run_id = opportunity_watch_service.create_watch_sync_run(db, "ws_local").id
    finally:
        db.close()

    opportunity_watch_service.execute_sync_run(
        run_id,
        transport=_mock_transport(handler),
        sleep_fn=lambda _s: None,
    )

    db = SessionLocal()
    try:
        hits = opportunity_watch_service.list_watch_hits(db, "ws_local")
        assert len(hits) == 2
        assert {item.watch_plan_id for item in hits} == {
            "watch_plan_sync_1",
            "watch_plan_sync_2",
        }
        assert {item.source_info_id for item in hits} == {_INFO_ID_A}
        assert detail_hits["count"] == 1
        run = opportunity_watch_service.get_watch_sync_run(db, "ws_local", run_id)
        assert run is not None
        assert run.detail_page_count == 1
        assert run.status == "succeeded"
    finally:
        db.close()


def test_sync_execute_detail_cap_fifty_and_preserves_existing():
    """用途：详情最多 50；失败路径不删除既有命中与本地标讯。"""
    _seed_enabled_plan(title="大量候选计划")
    db = SessionLocal()
    try:
        existing_run = BidSourceSyncRunRow(
            id="watch_run_existing",
            workspace_id="ws_local",
            source_name="chnenergy",
            status="succeeded",
            started_at=datetime(2026, 7, 1, tzinfo=timezone.utc),
            finished_at=datetime(2026, 7, 1, 1, tzinfo=timezone.utc),
        )
        db.add(existing_run)
        db.flush()
        db.add(
            BidSourceHitRow(
                id="watch_hit_existing",
                workspace_id="ws_local",
                watch_plan_id="watch_plan_sync_a",
                sync_run_id=existing_run.id,
                source_name="chnenergy",
                source_info_id="aaaaaaaa-bbbb-cccc-dddd-eeeeeeeeeeee",
                category_num="001002001",
                source_publish_text="2026-07-01 00:00:00",
                title="既有命中",
                deadline_at_local="2026-07-10 09:00:00",
                opening_at_local="2026-07-10 09:00:00",
                source_timezone="Asia/Shanghai",
                extraction_status="resolved",
            )
        )
        db.add(
            BidOpportunityRow(
                id="opp_existing_sync",
                workspace_id="ws_local",
                title="既有本地标讯",
                buyer="招标人",
                region="北京",
                budget_label="",
                deadline=date(2026, 8, 1),
                tags_json=None,
                summary="",
                source_label="本地录入",
                source_key="manual:keep",
            )
        )
        # 再写入 12 个计划，每计划 5 条唯一候选 → 60，详情封顶 50
        for i in range(12):
            db.add(
                BidWatchPlanRow(
                    id=f"watch_plan_cap_{i}",
                    workspace_id="ws_local",
                    title=f"计划{i}",
                    buyer="b",
                    scope="s",
                    duration="",
                    expected_publish_text="",
                    remark="",
                    fingerprint=f"fp-cap-{i}",
                    enabled=True,
                )
            )
        db.commit()
    finally:
        db.close()

    detail_hits = {"count": 0}
    search_calls = {"n": 0}

    def _uuid(i: int) -> str:
        return f"b2363623-ea1e-4cc1-8e2d-{i:012d}"

    def handler(request: httpx.Request) -> httpx.Response:
        url = str(request.url)
        if request.method == "GET" and url.rstrip("/") == _PORTAL_URL.rstrip("/"):
            return httpx.Response(
                200,
                headers=[("set-cookie", "uid=ok; Path=/")],
                text="portal",
            )
        if request.method == "POST":
            search_calls["n"] += 1
            base = (search_calls["n"] - 1) * 5
            return httpx.Response(
                200,
                json={
                    "result": {
                        "records": [
                            {
                                "title": f"公告{base + i}",
                                "infodate": "2026-07-09 10:00:00",
                                "linkurl": _jump(_uuid(base + i)),
                            }
                            for i in range(5)
                        ]
                    }
                },
            )
        if request.method == "GET" and url.endswith(".html"):
            detail_hits["count"] += 1
            return httpx.Response(200, text=_deadline_html())
        raise AssertionError(url)

    db = SessionLocal()
    try:
        run_id = opportunity_watch_service.create_watch_sync_run(db, "ws_local").id
    finally:
        db.close()

    opportunity_watch_service.execute_sync_run(
        run_id,
        transport=_mock_transport(handler),
        sleep_fn=lambda _s: None,
    )

    db = SessionLocal()
    try:
        run = opportunity_watch_service.get_watch_sync_run(db, "ws_local", run_id)
        assert run is not None
        assert run.detail_page_count == 50
        assert detail_hits["count"] == 50
        assert run.status == "partial"
        assert db.get(BidSourceHitRow, "watch_hit_existing") is not None
        assert db.get(BidOpportunityRow, "opp_existing_sync") is not None
    finally:
        db.close()


def test_sync_execute_stops_on_two_consecutive_network_errors():
    _seed_enabled_plan(title="网络失败计划")

    def handler(request: httpx.Request) -> httpx.Response:
        if request.method == "GET" and str(request.url).rstrip("/") == _PORTAL_URL.rstrip(
            "/"
        ):
            return httpx.Response(
                200,
                headers=[("set-cookie", "uid=ok; Path=/")],
                text="portal",
            )
        raise httpx.ConnectError("simulated", request=request)

    db = SessionLocal()
    try:
        run_id = opportunity_watch_service.create_watch_sync_run(db, "ws_local").id
    finally:
        db.close()

    # 默认检索重试 1 次：两次连接失败触发连续网络失败停止
    opportunity_watch_service.execute_sync_run(
        run_id,
        transport=_mock_transport(handler),
        sleep_fn=lambda _s: None,
    )
    db = SessionLocal()
    try:
        run = opportunity_watch_service.get_watch_sync_run(db, "ws_local", run_id)
        assert run is not None
        assert run.status == "failed"
        assert run.error_code == "source_unavailable"
        assert run.finished_at is not None
    finally:
        db.close()


def test_sync_api_post_and_get_run_status(client: TestClient, monkeypatch: pytest.MonkeyPatch):
    """用途：POST /sync 202+runId；GET 轮询；并发 409；跨空间 404。"""
    _seed_enabled_plan(title="接口同步计划")

    def handler(request: httpx.Request) -> httpx.Response:
        url = str(request.url)
        if request.method == "GET" and url.rstrip("/") == _PORTAL_URL.rstrip("/"):
            return httpx.Response(
                200,
                headers=[("set-cookie", "uid=ok; Path=/")],
                text="portal",
            )
        if request.method == "POST":
            return httpx.Response(
                200,
                json={
                    "result": {
                        "records": [
                            {
                                "title": "接口公告",
                                "infodate": "2026-07-09 17:14:11",
                                "linkurl": _jump(_INFO_ID_A),
                            }
                        ]
                    }
                },
            )
        if request.method == "GET" and url.endswith(".html"):
            return httpx.Response(200, text=_deadline_html())
        raise AssertionError(url)

    real_execute = opportunity_watch_service.execute_sync_run

    def _patched_execute(run_id: str, **kwargs):
        kwargs.setdefault("transport", _mock_transport(handler))
        kwargs.setdefault("sleep_fn", lambda _s: None)
        return real_execute(run_id, **kwargs)

    monkeypatch.setattr(opportunity_watch_service, "execute_sync_run", _patched_execute)

    resp = client.post("/api/opportunity-watch/sync")
    assert resp.status_code == 202
    body = resp.json()
    assert set(body.keys()) == {"runId"}
    run_id = body["runId"]
    assert run_id
    assert OpportunityWatchSyncAcceptedOut.model_validate(
        {"runId": run_id}
    ).run_id == run_id

    got = client.get(f"/api/opportunity-watch/runs/{run_id}")
    assert got.status_code == 200
    payload = got.json()
    assert payload["id"] == run_id
    assert payload["status"] == "succeeded"
    assert payload["sourceName"] == "chnenergy"
    assert payload["resolvedCount"] == 1
    assert "cookie" not in json.dumps(payload).lower()

    db = SessionLocal()
    try:
        db.add(
            BidSourceSyncRunRow(
                id="watch_run_block",
                workspace_id="ws_local",
                source_name="chnenergy",
                status="running",
                started_at=datetime.now(timezone.utc),
            )
        )
        db.commit()
    finally:
        db.close()
    conflict = client.post("/api/opportunity-watch/sync")
    assert conflict.status_code == 409

    other = "ws_sync_api_other"
    _create_watch_workspace(other)
    missing = client.get(
        f"/api/opportunity-watch/runs/{run_id}",
        headers={"X-Workspace-Id": other},
    )
    assert missing.status_code == 404

def _seed_accept_hit(
    *,
    hit_id: str = "watch_hit_accept_a",
    workspace_id: str = "ws_local",
    plan_id: str = "watch_plan_accept_a",
    run_id: str = "watch_run_accept_a",
    source_info_id: str = "b2363623-ea1e-4cc1-8e2d-0c2d2850b697",
    title: str = "某项目招标公告（接受）",
    extraction_status: str = "resolved",
    deadline_at_local: str | None = "2026-07-29 09:00:00",
    buyer: str = "国能招标人甲",
    scope: str = "甲供范围摘要",
) -> None:
    """用途：为人工接受用例准备同空间计划、运行与命中。"""
    db = SessionLocal()
    try:
        if db.get(BidWatchPlanRow, plan_id) is None:
            db.add(
                BidWatchPlanRow(
                    id=plan_id,
                    workspace_id=workspace_id,
                    title="接受测试计划",
                    buyer=buyer,
                    scope=scope,
                    duration="",
                    expected_publish_text="",
                    remark="",
                    fingerprint=f"fp-accept-{plan_id}",
                    enabled=True,
                )
            )
        if db.get(BidSourceSyncRunRow, run_id) is None:
            db.add(
                BidSourceSyncRunRow(
                    id=run_id,
                    workspace_id=workspace_id,
                    source_name="chnenergy",
                    status="succeeded",
                    started_at=datetime(2026, 7, 13, 8, tzinfo=timezone.utc),
                    finished_at=datetime(2026, 7, 13, 8, 5, tzinfo=timezone.utc),
                )
            )
        db.flush()
        db.add(
            BidSourceHitRow(
                id=hit_id,
                workspace_id=workspace_id,
                watch_plan_id=plan_id,
                sync_run_id=run_id,
                source_name="chnenergy",
                source_info_id=source_info_id,
                category_num="001002001",
                source_publish_text="2026-07-09 17:14:11",
                title=title,
                deadline_at_local=deadline_at_local,
                opening_at_local=deadline_at_local,
                source_timezone="Asia/Shanghai",
                extraction_status=extraction_status,
            )
        )
        db.commit()
    finally:
        db.close()


def test_accept_resolved_hit_creates_local_opportunity_fields():
    """用途：resolved 命中接受后创建本地标讯，字段与 source_key 符合冻结契约。"""
    _seed_accept_hit()
    db = SessionLocal()
    try:
        result = opportunity_watch_service.accept_watch_hit(
            db, "ws_local", "watch_hit_accept_a"
        )
        assert result["created"] is True
        opp_id = result["opportunity_id"]
        opp = db.get(BidOpportunityRow, opp_id)
        assert opp is not None
        assert opp.workspace_id == "ws_local"
        assert opp.title == "某项目招标公告（接受）"
        assert opp.buyer == "国能招标人甲"
        assert opp.summary == "甲供范围摘要"
        assert opp.region == "其他"
        assert opp.source_label == "国能 e 招计划追踪"
        assert opp.deadline == date(2026, 7, 29)
        assert opp.source_key == "chnenergy:b2363623-ea1e-4cc1-8e2d-0c2d2850b697"
        hit = db.get(BidSourceHitRow, "watch_hit_accept_a")
        assert hit is not None
        assert hit.accepted_opportunity_id == opp_id
    finally:
        db.close()


def test_accept_needs_review_or_missing_deadline_rejected():
    """用途：needs_review 或缺截止时间为 400 等价服务校验错误。"""
    _seed_accept_hit(
        hit_id="watch_hit_accept_review",
        plan_id="watch_plan_accept_review",
        run_id="watch_run_accept_review",
        source_info_id="aaaaaaaa-bbbb-cccc-dddd-111111111111",
        extraction_status="needs_review",
        deadline_at_local=None,
    )
    _seed_accept_hit(
        hit_id="watch_hit_accept_notime",
        plan_id="watch_plan_accept_notime",
        run_id="watch_run_accept_notime",
        source_info_id="aaaaaaaa-bbbb-cccc-dddd-222222222222",
        extraction_status="resolved",
        deadline_at_local="",
    )
    _seed_accept_hit(
        hit_id="watch_hit_accept_baddate",
        plan_id="watch_plan_accept_baddate",
        run_id="watch_run_accept_baddate",
        source_info_id="aaaaaaaa-bbbb-cccc-dddd-333333333333",
        extraction_status="resolved",
        deadline_at_local="not-a-date",
    )
    db = SessionLocal()
    try:
        with pytest.raises(opportunity_watch_service.WatchHitAcceptValidationError):
            opportunity_watch_service.accept_watch_hit(
                db, "ws_local", "watch_hit_accept_review"
            )
        with pytest.raises(opportunity_watch_service.WatchHitAcceptValidationError):
            opportunity_watch_service.accept_watch_hit(
                db, "ws_local", "watch_hit_accept_notime"
            )
        with pytest.raises(opportunity_watch_service.WatchHitAcceptValidationError):
            opportunity_watch_service.accept_watch_hit(
                db, "ws_local", "watch_hit_accept_baddate"
            )
        assert db.get(BidSourceHitRow, "watch_hit_accept_review").accepted_opportunity_id is None
        assert (
            db.scalars(
                select(BidOpportunityRow).where(
                    BidOpportunityRow.source_key
                    == "chnenergy:aaaaaaaa-bbbb-cccc-dddd-111111111111"
                )
            ).first()
            is None
        )
    finally:
        db.close()


def test_accept_is_idempotent_and_reuses_source_key():
    """用途：重复接受与同空间既有 source_key 均复用标讯，第二次 created=False。"""
    info_id = "cccccccc-dddd-eeee-ffff-000000000001"
    _seed_accept_hit(
        hit_id="watch_hit_accept_idem",
        plan_id="watch_plan_accept_idem",
        run_id="watch_run_accept_idem",
        source_info_id=info_id,
        title="幂等接受公告",
    )
    first_opp_id: str
    db = SessionLocal()
    try:
        first = opportunity_watch_service.accept_watch_hit(
            db, "ws_local", "watch_hit_accept_idem"
        )
        second = opportunity_watch_service.accept_watch_hit(
            db, "ws_local", "watch_hit_accept_idem"
        )
        assert first["created"] is True
        assert second["created"] is False
        assert first["opportunity_id"] == second["opportunity_id"]
        first_opp_id = first["opportunity_id"]
        count = len(
            list(
                db.scalars(
                    select(BidOpportunityRow).where(
                        BidOpportunityRow.workspace_id == "ws_local",
                        BidOpportunityRow.source_key == f"chnenergy:{info_id}",
                    )
                ).all()
            )
        )
        assert count == 1
    finally:
        db.close()

    # 另一命中同 info_id：复用既有 source_key 标讯
    _seed_accept_hit(
        hit_id="watch_hit_accept_reuse",
        plan_id="watch_plan_accept_reuse",
        run_id="watch_run_accept_reuse",
        source_info_id=info_id,
        title="另一计划下的同公告",
        buyer="另一招标人",
        scope="另一范围",
    )
    db = SessionLocal()
    try:
        reused = opportunity_watch_service.accept_watch_hit(
            db, "ws_local", "watch_hit_accept_reuse"
        )
        assert reused["created"] is False
        assert reused["opportunity_id"] == first_opp_id
        hit = db.get(BidSourceHitRow, "watch_hit_accept_reuse")
        assert hit is not None
        assert hit.accepted_opportunity_id == first_opp_id
    finally:
        db.close()


def test_accept_cross_workspace_not_found_and_no_partial_on_failure():
    """用途：跨空间不可读取；失败路径不留下半成品标讯或回写。"""
    other = "ws_accept_other"
    _create_watch_workspace(other)
    _seed_accept_hit(
        hit_id="watch_hit_accept_other",
        workspace_id=other,
        plan_id="watch_plan_accept_other",
        run_id="watch_run_accept_other",
        source_info_id="dddddddd-eeee-ffff-aaaa-111111111111",
    )
    db = SessionLocal()
    try:
        with pytest.raises(opportunity_watch_service.WatchHitNotFoundError):
            opportunity_watch_service.accept_watch_hit(
                db, "ws_local", "watch_hit_accept_other"
            )
        with pytest.raises(opportunity_watch_service.WatchHitNotFoundError):
            opportunity_watch_service.accept_watch_hit(
                db, "ws_local", "watch_hit_missing"
            )
    finally:
        db.close()

    _seed_accept_hit(
        hit_id="watch_hit_accept_fail",
        plan_id="watch_plan_accept_fail",
        run_id="watch_run_accept_fail",
        source_info_id="eeeeeeee-ffff-aaaa-bbbb-222222222222",
        title="失败回滚样本",
    )
    db = SessionLocal()
    try:
        hit = db.get(BidSourceHitRow, "watch_hit_accept_fail")
        assert hit is not None
        # 模拟事务中断：校验失败不得创建半成品
        hit.extraction_status = "needs_review"
        db.commit()
        with pytest.raises(opportunity_watch_service.WatchHitAcceptValidationError):
            opportunity_watch_service.accept_watch_hit(
                db, "ws_local", "watch_hit_accept_fail"
            )
        db.refresh(hit)
        assert hit.accepted_opportunity_id is None
        assert (
            db.scalars(
                select(BidOpportunityRow).where(
                    BidOpportunityRow.source_key
                    == "chnenergy:eeeeeeee-ffff-aaaa-bbbb-222222222222"
                )
            ).first()
            is None
        )
    finally:
        db.close()


def test_accept_past_deadline_still_creates_opportunity():
    """用途：已截止命中仍可写入本地标讯，不在此拦截立项。"""
    _seed_accept_hit(
        hit_id="watch_hit_accept_past",
        plan_id="watch_plan_accept_past",
        run_id="watch_run_accept_past",
        source_info_id="ffffffff-aaaa-bbbb-cccc-333333333333",
        deadline_at_local="2020-01-02 09:00:00",
        title="已截止公告",
    )
    db = SessionLocal()
    try:
        result = opportunity_watch_service.accept_watch_hit(
            db, "ws_local", "watch_hit_accept_past"
        )
        assert result["created"] is True
        opp = db.get(BidOpportunityRow, result["opportunity_id"])
        assert opp is not None
        assert opp.deadline == date(2020, 1, 2)
    finally:
        db.close()


def test_accept_api_status_codes_and_camel_case(client: TestClient):
    """用途：POST accept 固定状态码与 camelCase 响应，无请求体字段。"""
    _seed_accept_hit(
        hit_id="watch_hit_accept_api",
        plan_id="watch_plan_accept_api",
        run_id="watch_run_accept_api",
        source_info_id="aaaaaaaa-bbbb-cccc-dddd-api000000001",
        title="接口接受公告",
    )
    resp = client.post("/api/opportunity-watch/hits/watch_hit_accept_api/accept")
    assert resp.status_code == 200
    body = resp.json()
    assert set(body.keys()) == {"opportunityId", "created"}
    assert body["created"] is True
    assert body["opportunityId"]
    assert OpportunityWatchAcceptOut.model_validate(
        {"opportunityId": body["opportunityId"], "created": True}
    ).opportunity_id == body["opportunityId"]

    again = client.post("/api/opportunity-watch/hits/watch_hit_accept_api/accept")
    assert again.status_code == 200
    again_body = again.json()
    assert again_body["created"] is False
    assert again_body["opportunityId"] == body["opportunityId"]

    _seed_accept_hit(
        hit_id="watch_hit_accept_api_review",
        plan_id="watch_plan_accept_api_r",
        run_id="watch_run_accept_api_r",
        source_info_id="aaaaaaaa-bbbb-cccc-dddd-api000000002",
        extraction_status="needs_review",
        deadline_at_local=None,
    )
    bad = client.post("/api/opportunity-watch/hits/watch_hit_accept_api_review/accept")
    assert bad.status_code == 400

    missing = client.post("/api/opportunity-watch/hits/watch_hit_accept_missing/accept")
    assert missing.status_code == 404

    other = "ws_accept_api_other"
    _create_watch_workspace(other)
    _seed_accept_hit(
        hit_id="watch_hit_accept_api_other",
        workspace_id=other,
        plan_id="watch_plan_accept_api_o",
        run_id="watch_run_accept_api_o",
        source_info_id="aaaaaaaa-bbbb-cccc-dddd-api000000003",
    )
    cross = client.post(
        "/api/opportunity-watch/hits/watch_hit_accept_api_other/accept",
        headers={"X-Workspace-Id": "ws_local"},
    )
    assert cross.status_code == 404
    # 响应不得泄漏 URL/Cookie/HTML
    for payload in (body, again_body, bad.json(), missing.json()):
        lowered = json.dumps(payload, ensure_ascii=False).lower()
        assert "cookie" not in lowered
        assert "http" not in lowered
        assert "<html" not in lowered


# ---------- P9B 任务6：dashboard 只读聚合 ----------


def _seed_dashboard_hit(
    *,
    hit_id: str,
    workspace_id: str = "ws_local",
    plan_id: str = "watch_plan_dash",
    run_id: str = "watch_run_dash",
    source_info_id: str = "b2363623-ea1e-4cc1-8e2d-0c2d2850b697",
    category_num: str = "001002001",
    source_publish_text: str = "2026-07-09 17:14:11",
    title: str = "仪表盘公告甲",
    extraction_status: str = "resolved",
    deadline_at_local: str | None = "2026-07-29 09:00:00",
    opening_at_local: str | None = "2026-07-29 09:00:00",
    updated_at: datetime | None = None,
) -> None:
    """用途：为 dashboard 只读用例准备计划/运行/命中。"""
    db = SessionLocal()
    try:
        if db.get(BidWatchPlanRow, plan_id) is None:
            db.add(
                BidWatchPlanRow(
                    id=plan_id,
                    workspace_id=workspace_id,
                    title="仪表盘计划",
                    buyer="招标人甲",
                    scope="范围甲",
                    duration="",
                    expected_publish_text="",
                    remark="",
                    fingerprint=f"fp-dash-{plan_id}",
                    enabled=True,
                )
            )
        if db.get(BidSourceSyncRunRow, run_id) is None:
            db.add(
                BidSourceSyncRunRow(
                    id=run_id,
                    workspace_id=workspace_id,
                    source_name="chnenergy",
                    status="succeeded",
                    started_at=datetime(2026, 7, 13, 8, tzinfo=timezone.utc),
                    finished_at=datetime(2026, 7, 13, 8, 5, tzinfo=timezone.utc),
                    plan_count=1,
                    candidate_count=1,
                    detail_page_count=1,
                    resolved_count=1,
                    needs_review_count=0,
                    skipped_count=0,
                )
            )
        db.flush()
        hit = BidSourceHitRow(
            id=hit_id,
            workspace_id=workspace_id,
            watch_plan_id=plan_id,
            sync_run_id=run_id,
            source_name="chnenergy",
            source_info_id=source_info_id,
            category_num=category_num,
            source_publish_text=source_publish_text,
            title=title,
            deadline_at_local=deadline_at_local,
            opening_at_local=opening_at_local,
            source_timezone="Asia/Shanghai",
            extraction_status=extraction_status,
        )
        if updated_at is not None:
            hit.updated_at = updated_at
        db.add(hit)
        db.commit()
    finally:
        db.close()


def test_dashboard_returns_plan_count_latest_run_and_hits_desc(
    client: TestClient,
):
    """用途：当前空间返回计划数、最近运行、命中按更新时间倒序，且含动态 announcementUrl。"""
    older = datetime(2026, 7, 13, 9, 0, tzinfo=timezone.utc)
    newer = datetime(2026, 7, 13, 10, 0, tzinfo=timezone.utc)
    _seed_dashboard_hit(
        hit_id="watch_hit_dash_old",
        source_info_id="aaaaaaaa-bbbb-cccc-dddd-000000000001",
        title="较旧命中",
        updated_at=older,
    )
    _seed_dashboard_hit(
        hit_id="watch_hit_dash_new",
        plan_id="watch_plan_dash_b",
        run_id="watch_run_dash_b",
        source_info_id="aaaaaaaa-bbbb-cccc-dddd-000000000002",
        title="较新命中",
        updated_at=newer,
    )
    # 另一计划仅增加 planCount
    db = SessionLocal()
    try:
        db.add(
            BidWatchPlanRow(
                id="watch_plan_dash_extra",
                workspace_id="ws_local",
                title="额外计划",
                buyer="",
                scope="",
                duration="",
                expected_publish_text="",
                remark="",
                fingerprint="fp-dash-extra",
                enabled=True,
            )
        )
        db.commit()
    finally:
        db.close()

    resp = client.get("/api/opportunity-watch/dashboard")
    assert resp.status_code == 200
    body = resp.json()
    assert set(body.keys()) == {"planCount", "latestRun", "hits"}
    assert body["planCount"] == 3
    assert body["latestRun"] is not None
    assert body["latestRun"]["sourceName"] == "chnenergy"
    assert "cookie" not in json.dumps(body["latestRun"]).lower()

    titles = [item["title"] for item in body["hits"]]
    assert titles[0] == "较新命中"
    assert titles[1] == "较旧命中"
    assert len(body["hits"]) == 2

    hit = body["hits"][0]
    assert hit["extractionStatus"] == "resolved"
    assert hit["deadlineAtLocal"] == "2026-07-29 09:00:00"
    assert hit["openingAtLocal"] == "2026-07-29 09:00:00"
    assert hit["sourceTimezone"] == "Asia/Shanghai"
    assert hit["announcementUrl"] == (
        "https://www.chnenergybidding.com.cn/bidweb/"
        "001/001002/001002001/20260709/"
        "aaaaaaaa-bbbb-cccc-dddd-000000000002.html"
    )
    # 仅结构化字段 + 动态链接；不得回传 HTML/Cookie/原文
    lowered = json.dumps(hit, ensure_ascii=False).lower()
    assert "cookie" not in lowered
    assert "<html" not in lowered
    assert "raw" not in lowered


def test_dashboard_isolates_workspace_and_skips_invalid_announcement_url(
    client: TestClient,
):
    """用途：跨空间不可见；非法详情字段不生成 announcementUrl。"""
    other = "ws_dash_other"
    _create_watch_workspace(other)
    _seed_dashboard_hit(
        hit_id="watch_hit_dash_local",
        source_info_id="bbbbbbbb-cccc-dddd-eeee-000000000011",
        title="本空间命中",
    )
    _seed_dashboard_hit(
        hit_id="watch_hit_dash_other",
        workspace_id=other,
        plan_id="watch_plan_dash_other",
        run_id="watch_run_dash_other",
        source_info_id="bbbbbbbb-cccc-dddd-eeee-000000000012",
        title="他空间命中",
    )
    _seed_dashboard_hit(
        hit_id="watch_hit_dash_bad_fields",
        plan_id="watch_plan_dash_bad",
        run_id="watch_run_dash_bad",
        source_info_id="not-a-uuid",
        category_num="001006001",
        source_publish_text="not-a-date",
        title="非法字段命中",
        extraction_status="needs_review",
        deadline_at_local=None,
        opening_at_local=None,
    )

    local = client.get("/api/opportunity-watch/dashboard")
    assert local.status_code == 200
    body = local.json()
    titles = {item["title"] for item in body["hits"]}
    assert "本空间命中" in titles
    assert "他空间命中" not in titles
    assert "非法字段命中" in titles
    bad = next(item for item in body["hits"] if item["title"] == "非法字段命中")
    assert bad["announcementUrl"] is None
    assert bad["extractionStatus"] == "needs_review"

    other_resp = client.get(
        "/api/opportunity-watch/dashboard",
        headers={"X-Workspace-Id": other},
    )
    assert other_resp.status_code == 200
    other_body = other_resp.json()
    assert other_body["planCount"] == 1
    assert [item["title"] for item in other_body["hits"]] == ["他空间命中"]


def test_dashboard_is_read_only_without_sync_or_accept_side_effects(
    client: TestClient,
):
    """用途：GET dashboard 不触发同步、不写入命中、不创建本地标讯。"""
    _seed_dashboard_hit(
        hit_id="watch_hit_dash_readonly",
        source_info_id="cccccccc-dddd-eeee-ffff-000000000021",
        title="只读命中",
    )
    before_hits = SessionLocal()
    try:
        hit_count = len(
            list(
                before_hits.scalars(
                    select(BidSourceHitRow).where(
                        BidSourceHitRow.workspace_id == "ws_local"
                    )
                ).all()
            )
        )
        opp_count = len(
            list(
                before_hits.scalars(
                    select(BidOpportunityRow).where(
                        BidOpportunityRow.workspace_id == "ws_local"
                    )
                ).all()
            )
        )
        run_count = len(
            list(
                before_hits.scalars(
                    select(BidSourceSyncRunRow).where(
                        BidSourceSyncRunRow.workspace_id == "ws_local"
                    )
                ).all()
            )
        )
    finally:
        before_hits.close()

    resp = client.get("/api/opportunity-watch/dashboard")
    assert resp.status_code == 200

    after = SessionLocal()
    try:
        assert (
            len(
                list(
                    after.scalars(
                        select(BidSourceHitRow).where(
                            BidSourceHitRow.workspace_id == "ws_local"
                        )
                    ).all()
                )
            )
            == hit_count
        )
        assert (
            len(
                list(
                    after.scalars(
                        select(BidOpportunityRow).where(
                            BidOpportunityRow.workspace_id == "ws_local"
                        )
                    ).all()
                )
            )
            == opp_count
        )
        assert (
            len(
                list(
                    after.scalars(
                        select(BidSourceSyncRunRow).where(
                            BidSourceSyncRunRow.workspace_id == "ws_local"
                        )
                    ).all()
                )
            )
            == run_count
        )
        hit = after.get(BidSourceHitRow, "watch_hit_dash_readonly")
        assert hit is not None
        assert hit.accepted_opportunity_id is None
    finally:
        after.close()
