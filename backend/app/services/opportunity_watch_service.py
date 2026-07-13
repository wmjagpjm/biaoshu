"""
模块：国能 e 招计划追踪服务
用途：提供工作空间隔离的计划/运行/命中读取、中断运行收敛、本机 .xlsx 计划导入、受控同步，以及命中人工接受。
对接：app.main.lifespan；app.api.opportunity_watch；chnenergy_client；BidWatchPlanRow/BidOpportunityRow。
二次开发：不得在此混入任意 URL、Cookie、HTML、JSON 或浏览器请求；外部访问仅能由固定来源客户端承担；接受不得自动立项。
"""

from __future__ import annotations

import hashlib
import io
import re
import secrets
from datetime import date, datetime, timezone
from typing import Any, Callable

from openpyxl import load_workbook
from sqlalchemy import select, update
from sqlalchemy.orm import Session

from app.core.config import get_settings
from app.core.database import SessionLocal
from app.models.entities import (
    BidOpportunityRow,
    BidSourceHitRow,
    BidSourceSyncRunRow,
    BidWatchPlanRow,
)
from app.services.chnenergy_client import (
    ChnenergyClientError,
    ChnenergyControlledClient,
    ChnenergyNetworkError,
    ChnenergySyncStopError,
    build_notice_detail_url,
    extract_notice_times,
    parse_jump_fields,
)

# 人工接受写入本地标讯时的固定展示字段；不得改成 URL 或远端原文。
_ACCEPT_REGION = "其他"
_ACCEPT_SOURCE_LABEL = "国能 e 招计划追踪"
_ACCEPT_SOURCE_KEY_PREFIX = "chnenergy:"
# 完整北京时间：YYYY-MM-DD HH:mm:ss
_DEADLINE_LOCAL_RE = re.compile(
    r"^(\d{4}-\d{2}-\d{2})[ T](\d{2}:\d{2}:\d{2})$"
)

# 计划表中文表头到实体字段的固定映射；未知列忽略。
_PLAN_HEADER_TO_FIELD = {
    "招标计划名称": "title",
    "招标人": "buyer",
    "范围": "scope",
    "计划工期": "duration",
    "预计发布公告时间": "expected_publish_text",
    "备注": "remark",
}
_REQUIRED_TITLE_HEADER = "招标计划名称"
_HEADER_SCAN_MAX_ROW = 10
_FIELD_LIMITS = {
    "title": 500,
    "buyer": 500,
    "scope": 20000,
    "duration": 300,
    "expected_publish_text": 300,
    "remark": 20000,
}


class WatchPlanImportValidationError(ValueError):
    """
    模块：计划表导入校验异常
    用途：承载可定位的行号/字段错误，保证路由返回 422 且整批不写入。
    对接：import_watch_plans_from_xlsx；POST /api/opportunity-watch/plans/import。
    二次开发：errors 仅允许安全的 row/field/message；禁止拼接原始文件、路径、公式或 Python 异常原文。
    """

    def __init__(self, errors: list[dict[str, Any]]):
        super().__init__("计划表导入数据不合法")
        self.errors = errors


class WatchSyncConflictError(ValueError):
    """
    模块：同步运行并发冲突
    用途：同工作空间已存在 queued/running 运行时拒绝新建。
    对接：create_watch_sync_run；POST /api/opportunity-watch/sync → 409。
    """

    def __init__(self) -> None:
        super().__init__("当前工作空间已有进行中的同步运行")


class WatchHitNotFoundError(LookupError):
    """
    模块：命中不存在或不属于当前工作空间
    用途：人工接受时将跨空间/未知命中映射为 HTTP 404。
    对接：accept_watch_hit；POST /api/opportunity-watch/hits/{hit_id}/accept。
    二次开发：禁止区分“存在但属其它空间”与“不存在”，避免工作空间枚举。
    """

    def __init__(self) -> None:
        super().__init__("公告命中不存在")


class WatchHitAcceptValidationError(ValueError):
    """
    模块：命中人工接受校验失败
    用途：未解析、缺截止时间或时间非法时拒绝接受，映射为 HTTP 400。
    对接：accept_watch_hit；POST /api/opportunity-watch/hits/{hit_id}/accept。
    二次开发：message 仅允许安全中文说明，禁止拼接 URL、HTML 或远端错误原文。
    """

    def __init__(self, message: str = "仅可接受已解析完整截止时间的命中") -> None:
        super().__init__(message)


def _now() -> datetime:
    """用途：统一同步运行恢复与计划导入写入的 UTC 时间源。"""
    return datetime.now(timezone.utc)


def _new_watch_plan_id() -> str:
    """用途：生成服务端追踪计划主键，避免客户端指定持久化标识。"""
    return f"watch_plan_{secrets.token_hex(8)}"


def _new_sync_run_id() -> str:
    """用途：生成同步运行主键。"""
    return f"watch_run_{secrets.token_hex(8)}"


def _new_hit_id() -> str:
    """用途：生成公告命中主键。"""
    return f"watch_hit_{secrets.token_hex(8)}"


def _new_opportunity_id() -> str:
    """用途：生成人工接受后的本地标讯主键，不依赖既有 opportunity_service。"""
    return f"opp_{secrets.token_hex(8)}"


def _parse_deadline_date_local(deadline_at_local: str | None) -> date:
    """
    用途：从北京时间完整本地字符串严格解析日期部分，供写入 bid_opportunities.deadline。
    对接：accept_watch_hit；BidSourceHitRow.deadline_at_local。
    二次开发：仅接受 YYYY-MM-DD HH:mm:ss；非法日历日必须失败，不得静默回退。
    """
    text = (deadline_at_local or "").strip()
    match = _DEADLINE_LOCAL_RE.fullmatch(text)
    if match is None:
        raise WatchHitAcceptValidationError("命中截止时间缺失或格式非法")
    try:
        # 先校验完整时间，再取日期，避免 02-30 等非法日历被截断放过。
        datetime.strptime(text.replace("T", " "), "%Y-%m-%d %H:%M:%S")
        return date.fromisoformat(match.group(1))
    except ValueError as exc:
        raise WatchHitAcceptValidationError("命中截止时间缺失或格式非法") from exc


def _chnenergy_source_key(source_info_id: str) -> str:
    """用途：由公告 infoid 生成不透明本地来源键，禁止写入 URL 或 Cookie。"""
    return f"{_ACCEPT_SOURCE_KEY_PREFIX}{source_info_id}"


def _clean_text(value: Any, *, limit: int = 1000) -> str:
    """用途：清洗单元格文本，统一 strip 与截断，不保留公式对象。"""
    if value is None:
        return ""
    if isinstance(value, datetime):
        text = value.isoformat(sep=" ", timespec="seconds")
    elif isinstance(value, float) and value.is_integer():
        text = str(int(value))
    else:
        text = str(value)
    return text.strip()[:limit]


def _plan_fingerprint(title: str, buyer: str, scope: str) -> str:
    """
    用途：对清洗后的计划名、招标人、范围计算确定性指纹，供工作空间内幂等去重。
    对接：BidWatchPlanRow.fingerprint；(workspace_id, fingerprint) 唯一约束。
    """
    material = f"{title}\n{buyer}\n{scope}".encode("utf-8")
    return hashlib.sha256(material).hexdigest()


def _cell_values(row: tuple[Any, ...] | list[Any]) -> list[Any]:
    """用途：将 openpyxl 行单元格规范为原始值列表。"""
    values: list[Any] = []
    for cell in row:
        if hasattr(cell, "value"):
            values.append(cell.value)
        else:
            values.append(cell)
    return values


def _is_blank_row(values: list[Any]) -> bool:
    """用途：判断数据行是否全空，供空行跳过裁定使用。"""
    return all(_clean_text(value) == "" for value in values)


def _locate_header(rows: list[tuple[int, list[Any]]]) -> tuple[int, dict[str, int]]:
    """
    用途：在前十个 Excel 行定位中文表头，并建立字段到 0 基列下标的映射。
    对接：import_watch_plans_from_xlsx 的表头扫描边界。
    """
    for excel_row, values in rows:
        if excel_row > _HEADER_SCAN_MAX_ROW:
            break
        cleaned = [_clean_text(value, limit=100) for value in values]
        if _REQUIRED_TITLE_HEADER not in cleaned:
            continue
        mapping: dict[str, int] = {}
        for index, header in enumerate(cleaned):
            field = _PLAN_HEADER_TO_FIELD.get(header)
            if field and field not in mapping:
                mapping[field] = index
        if "title" in mapping:
            return excel_row, mapping
    raise WatchPlanImportValidationError(
        [
            {
                "row": None,
                "field": _REQUIRED_TITLE_HEADER,
                "message": "前十行未找到表头「招标计划名称」",
            }
        ]
    )


def _read_field(values: list[Any], mapping: dict[str, int], field: str) -> str:
    """用途：按表头映射读取并清洗单个计划字段。"""
    index = mapping.get(field)
    if index is None or index >= len(values):
        return ""
    return _clean_text(values[index], limit=_FIELD_LIMITS[field])


def list_watch_plans(db: Session, workspace_id: str) -> list[BidWatchPlanRow]:
    """
    模块：计划追踪列表查询
    用途：只读取当前工作空间的计划追踪记录，避免服务层遗漏归属过滤。
    对接：BidWatchPlanRow；OpportunityWatchPlanOut；计划导入后的列表校验。
    二次开发：调用方必须传入已校验的 workspace_id；禁止省略 where 过滤或跨工作空间查询。
    """
    stmt = (
        select(BidWatchPlanRow)
        .where(BidWatchPlanRow.workspace_id == workspace_id)
        .order_by(BidWatchPlanRow.updated_at.desc(), BidWatchPlanRow.id.asc())
    )
    return list(db.scalars(stmt).all())


def list_watch_runs(db: Session, workspace_id: str) -> list[BidSourceSyncRunRow]:
    """
    模块：同步运行列表查询
    用途：只读取当前工作空间的同步运行记录，供后续运行状态和审计展示。
    对接：BidSourceSyncRunRow；OpportunityWatchSyncRunOut；当前尚无公开 HTTP 路由。
    二次开发：调用方必须传入已校验的 workspace_id；禁止跨工作空间查询或回传错误原文。
    """
    stmt = (
        select(BidSourceSyncRunRow)
        .where(BidSourceSyncRunRow.workspace_id == workspace_id)
        .order_by(BidSourceSyncRunRow.started_at.desc(), BidSourceSyncRunRow.id.asc())
    )
    return list(db.scalars(stmt).all())


def list_watch_hits(db: Session, workspace_id: str) -> list[BidSourceHitRow]:
    """
    模块：公告命中列表查询
    用途：只读取当前工作空间的公告命中，绝不跨工作空间返回追踪结果。
    对接：BidSourceHitRow；OpportunityWatchHitOut；当前尚无公开 HTTP 路由。
    二次开发：调用方必须传入已校验的 workspace_id；不得返回其它 workspace 命中或外部正文。
    """
    stmt = (
        select(BidSourceHitRow)
        .where(BidSourceHitRow.workspace_id == workspace_id)
        .order_by(BidSourceHitRow.updated_at.desc(), BidSourceHitRow.id.asc())
    )
    return list(db.scalars(stmt).all())


def accept_watch_hit(
    db: Session,
    workspace_id: str,
    hit_id: str,
) -> dict[str, Any]:
    """
    模块：人工接受国能公告命中
    用途：在单事务内将当前工作空间 resolved 且有完整截止时间的命中创建或幂等复用为本地标讯。
    对接：POST /api/opportunity-watch/hits/{hit_id}/accept；BidOpportunityRow；BidSourceHitRow。
    二次开发：禁止批量接受、同步后自动调用、创建项目；不得持久化 URL/Cookie/HTML/JSON/正文。
    """
    try:
        hit = db.scalars(
            select(BidSourceHitRow).where(
                BidSourceHitRow.id == hit_id,
                BidSourceHitRow.workspace_id == workspace_id,
            )
        ).first()
        if hit is None:
            raise WatchHitNotFoundError()

        if hit.extraction_status != "resolved":
            raise WatchHitAcceptValidationError("命中尚未解析出可接受的截止时间")

        deadline = _parse_deadline_date_local(hit.deadline_at_local)
        source_key = _chnenergy_source_key(hit.source_info_id)
        now = _now()

        # 已回写过的命中：直接复用，避免重复创建。
        if hit.accepted_opportunity_id:
            existing_linked = db.scalars(
                select(BidOpportunityRow).where(
                    BidOpportunityRow.id == hit.accepted_opportunity_id,
                    BidOpportunityRow.workspace_id == workspace_id,
                )
            ).first()
            if existing_linked is not None:
                return {
                    "opportunity_id": existing_linked.id,
                    "created": False,
                }

        # 同空间同来源键：幂等复用既有标讯并回写命中。
        existing_by_key = db.scalars(
            select(BidOpportunityRow).where(
                BidOpportunityRow.workspace_id == workspace_id,
                BidOpportunityRow.source_key == source_key,
            )
        ).first()
        if existing_by_key is not None:
            hit.accepted_opportunity_id = existing_by_key.id
            hit.updated_at = now
            db.commit()
            return {
                "opportunity_id": existing_by_key.id,
                "created": False,
            }

        plan = db.scalars(
            select(BidWatchPlanRow).where(
                BidWatchPlanRow.id == hit.watch_plan_id,
                BidWatchPlanRow.workspace_id == workspace_id,
            )
        ).first()
        buyer = (plan.buyer if plan is not None else "") or ""
        summary = (plan.scope if plan is not None else "") or ""

        opportunity = BidOpportunityRow(
            id=_new_opportunity_id(),
            workspace_id=workspace_id,
            title=(hit.title or "")[:500],
            buyer=buyer[:500],
            region=_ACCEPT_REGION,
            budget_label="",
            deadline=deadline,
            tags_json=None,
            summary=summary[:20000],
            source_label=_ACCEPT_SOURCE_LABEL,
            source_key=source_key,
            created_at=now,
            updated_at=now,
        )
        db.add(opportunity)
        hit.accepted_opportunity_id = opportunity.id
        hit.updated_at = now
        db.commit()
        return {
            "opportunity_id": opportunity.id,
            "created": True,
        }
    except Exception:
        db.rollback()
        raise


def mark_interrupted_watch_runs(db: Session) -> int:
    """
    模块：中断同步运行收敛
    用途：将进程重启前未结束的 queued/running 同步运行标为 failed/interrupted。
    对接：app.main.lifespan；仅修改运行状态，命中和本地标讯均保留。
    二次开发：不得删除命中、计划或 bid_opportunities；新增运行状态时同步扩展过滤条件与测试。
    """
    now = _now()
    result = db.execute(
        update(BidSourceSyncRunRow)
        .where(BidSourceSyncRunRow.status.in_(("queued", "running")))
        .values(
            status="failed",
            error_code="interrupted",
            finished_at=now,
            updated_at=now,
        )
    )
    db.commit()
    return int(result.rowcount or 0)


def import_watch_plans_from_xlsx(
    db: Session,
    workspace_id: str,
    *,
    filename: str,
    content: bytes,
    max_rows: int,
) -> dict[str, int]:
    """
    模块：国能计划表内存导入
    用途：仅在内存中解析本机 .xlsx，全量校验后于单一事务按指纹幂等写入追踪计划。
    对接：POST /api/opportunity-watch/plans/import；BidWatchPlanRow；Settings 行数上限。
    二次开发：禁止持久化 content/工作簿/路径；不得接受 URL、Cookie、Token 或发起外网请求。
    """
    suffix = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if suffix != "xlsx":
        raise ValueError("仅支持 .xlsx 招标计划文件")

    try:
        workbook = load_workbook(
            io.BytesIO(content),
            read_only=True,
            data_only=True,
        )
    except Exception as exc:
        # 不向外透传 openpyxl 或 zip 原始异常文本。
        raise ValueError("无法解析 .xlsx 计划文件") from exc

    try:
        sheet = workbook.worksheets[0] if workbook.worksheets else None
        if sheet is None:
            raise WatchPlanImportValidationError(
                [{"row": None, "field": _REQUIRED_TITLE_HEADER, "message": "工作簿没有可用工作表"}]
            )

        # 流式读取：表头至多扫描前 10 行；定位后逐行处理，禁止全表物化。
        row_iter = enumerate(sheet.iter_rows(), start=1)
        header_scan: list[tuple[int, list[Any]]] = []
        for excel_row, row in row_iter:
            values = _cell_values(row)
            header_scan.append((excel_row, values))
            # 一旦本行含必填表头即停止扫描，后续行改由数据流处理。
            cleaned = [_clean_text(value, limit=100) for value in values]
            if _REQUIRED_TITLE_HEADER in cleaned:
                break
            if excel_row >= _HEADER_SCAN_MAX_ROW:
                break

        header_row, mapping = _locate_header(header_scan)

        payloads: list[dict[str, str]] = []
        errors: list[dict[str, Any]] = []
        non_blank_plan_rows = 0

        def _consume_data_row(excel_row: int, values: list[Any]) -> None:
            """用途：流式消费单行数据；超限立即抛错，全空白行跳过。"""
            nonlocal non_blank_plan_rows
            if excel_row <= header_row:
                return
            if _is_blank_row(values):
                return
            non_blank_plan_rows += 1
            if non_blank_plan_rows > max_rows:
                raise ValueError(f"导入计划行数不能超过 {max_rows} 行")
            title = _read_field(values, mapping, "title")
            buyer = _read_field(values, mapping, "buyer")
            scope = _read_field(values, mapping, "scope")
            duration = _read_field(values, mapping, "duration")
            expected_publish_text = _read_field(values, mapping, "expected_publish_text")
            remark = _read_field(values, mapping, "remark")
            if not title:
                errors.append(
                    {
                        "row": excel_row,
                        "field": _REQUIRED_TITLE_HEADER,
                        "message": "招标计划名称不能为空",
                    }
                )
                return
            payloads.append(
                {
                    "title": title,
                    "buyer": buyer,
                    "scope": scope,
                    "duration": duration,
                    "expected_publish_text": expected_publish_text,
                    "remark": remark,
                    "fingerprint": _plan_fingerprint(title, buyer, scope),
                }
            )

        # 表头扫描窗口内、表头之后的行先消费（仅当表头未提前命中时可能存在）。
        for excel_row, values in header_scan:
            _consume_data_row(excel_row, values)

        # 继续流式读取剩余行；触发上限时不再请求后续行。
        for excel_row, row in row_iter:
            _consume_data_row(excel_row, _cell_values(row))

        if errors:
            raise WatchPlanImportValidationError(errors)

        fingerprints = {item["fingerprint"] for item in payloads}
        existing: set[str] = set()
        if fingerprints:
            existing = set(
                db.scalars(
                    select(BidWatchPlanRow.fingerprint).where(
                        BidWatchPlanRow.workspace_id == workspace_id,
                        BidWatchPlanRow.fingerprint.in_(fingerprints),
                    )
                ).all()
            )

        inserted = 0
        skipped = 0
        seen_in_batch: set[str] = set()
        now = _now()
        for payload in payloads:
            fingerprint = payload["fingerprint"]
            if fingerprint in existing or fingerprint in seen_in_batch:
                skipped += 1
                continue
            seen_in_batch.add(fingerprint)
            db.add(
                BidWatchPlanRow(
                    id=_new_watch_plan_id(),
                    workspace_id=workspace_id,
                    title=payload["title"],
                    buyer=payload["buyer"],
                    scope=payload["scope"],
                    duration=payload["duration"],
                    expected_publish_text=payload["expected_publish_text"],
                    remark=payload["remark"],
                    fingerprint=fingerprint,
                    enabled=True,
                    created_at=now,
                    updated_at=now,
                )
            )
            inserted += 1

        db.commit()
        return {
            "inserted": inserted,
            "skipped": skipped,
            "total": len(payloads),
        }
    except Exception:
        db.rollback()
        raise
    finally:
        workbook.close()


def _count_active_sync_runs(db: Session, workspace_id: str) -> int:
    """用途：统计工作空间内仍活跃的 queued/running 同步运行。"""
    stmt = select(BidSourceSyncRunRow.id).where(
        BidSourceSyncRunRow.workspace_id == workspace_id,
        BidSourceSyncRunRow.status.in_(("queued", "running")),
    )
    return len(list(db.scalars(stmt).all()))


def _list_enabled_plans(db: Session, workspace_id: str, *, limit: int) -> list[BidWatchPlanRow]:
    """用途：读取当前工作空间启用中的计划，供同步检索。"""
    stmt = (
        select(BidWatchPlanRow)
        .where(
            BidWatchPlanRow.workspace_id == workspace_id,
            BidWatchPlanRow.enabled.is_(True),
        )
        .order_by(BidWatchPlanRow.created_at.asc(), BidWatchPlanRow.id.asc())
        .limit(limit)
    )
    return list(db.scalars(stmt).all())


def create_watch_sync_run(db: Session, workspace_id: str) -> BidSourceSyncRunRow:
    """
    模块：创建受控同步运行
    用途：在无并发活跃运行时创建 queued 记录；仅统计当前空间启用计划数。
    对接：POST /api/opportunity-watch/sync；BidSourceSyncRunRow。
    二次开发：禁止接受 URL/Cookie/搜索条件；并发冲突必须由调用方映射为 409。
    """
    if _count_active_sync_runs(db, workspace_id) > 0:
        raise WatchSyncConflictError()

    settings = get_settings()
    plans = _list_enabled_plans(
        db,
        workspace_id,
        limit=settings.max_opportunity_watch_plans_per_sync,
    )
    now = _now()
    row = BidSourceSyncRunRow(
        id=_new_sync_run_id(),
        workspace_id=workspace_id,
        source_name="chnenergy",
        status="queued",
        started_at=now,
        finished_at=None,
        plan_count=len(plans),
        candidate_count=0,
        detail_page_count=0,
        resolved_count=0,
        needs_review_count=0,
        skipped_count=0,
        error_code=None,
        created_at=now,
        updated_at=now,
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


def get_watch_sync_run(
    db: Session,
    workspace_id: str,
    run_id: str,
) -> BidSourceSyncRunRow | None:
    """
    模块：同步运行按主键读取
    用途：仅返回当前工作空间的运行；跨空间视为不存在。
    对接：GET /api/opportunity-watch/runs/{run_id}；OpportunityWatchSyncRunOut。
    """
    stmt = select(BidSourceSyncRunRow).where(
        BidSourceSyncRunRow.id == run_id,
        BidSourceSyncRunRow.workspace_id == workspace_id,
    )
    return db.scalars(stmt).first()


def _finalize_run(
    db: Session,
    run: BidSourceSyncRunRow,
    *,
    status: str,
    error_code: str | None,
    counters: dict[str, int],
) -> None:
    """用途：写入终态、脱敏计数与结束时间；不记录 URL/正文。"""
    now = _now()
    run.status = status
    run.error_code = error_code
    run.plan_count = counters.get("plan_count", run.plan_count)
    run.candidate_count = counters.get("candidate_count", 0)
    run.detail_page_count = counters.get("detail_page_count", 0)
    run.resolved_count = counters.get("resolved_count", 0)
    run.needs_review_count = counters.get("needs_review_count", 0)
    run.skipped_count = counters.get("skipped_count", 0)
    run.finished_at = now
    run.updated_at = now
    db.commit()


def _upsert_hit(
    db: Session,
    *,
    workspace_id: str,
    watch_plan_id: str,
    sync_run_id: str,
    source_info_id: str,
    category_num: str,
    source_publish_text: str,
    title: str,
    deadline_at_local: str | None,
    opening_at_local: str | None,
    extraction_status: str,
) -> BidSourceHitRow:
    """
    用途：按 (workspace, plan, info_id) 幂等写入/更新命中，不删除既有行。
    对接：BidSourceHitRow 唯一约束；跨计划可各自保留同一公告。
    """
    existing = db.scalars(
        select(BidSourceHitRow).where(
            BidSourceHitRow.workspace_id == workspace_id,
            BidSourceHitRow.watch_plan_id == watch_plan_id,
            BidSourceHitRow.source_info_id == source_info_id,
        )
    ).first()
    now = _now()
    if existing is None:
        row = BidSourceHitRow(
            id=_new_hit_id(),
            workspace_id=workspace_id,
            watch_plan_id=watch_plan_id,
            sync_run_id=sync_run_id,
            source_name="chnenergy",
            source_info_id=source_info_id,
            category_num=category_num,
            source_publish_text=source_publish_text[:100],
            title=title[:1000],
            deadline_at_local=deadline_at_local,
            opening_at_local=opening_at_local,
            source_timezone="Asia/Shanghai",
            extraction_status=extraction_status,
            accepted_opportunity_id=None,
            created_at=now,
            updated_at=now,
        )
        db.add(row)
        return row

    existing.sync_run_id = sync_run_id
    existing.category_num = category_num
    existing.source_publish_text = source_publish_text[:100]
    existing.title = title[:1000]
    existing.deadline_at_local = deadline_at_local
    existing.opening_at_local = opening_at_local
    existing.extraction_status = extraction_status
    existing.updated_at = now
    return existing


def execute_sync_run(
    run_id: str,
    *,
    transport: Any | None = None,
    sleep_fn: Callable[[float], None] | None = None,
    min_interval_seconds: float | None = None,
    search_retry_count: int | None = None,
) -> None:
    """
    模块：执行国能 e 招受控同步
    用途：独立打开数据库会话，将 queued 运行推进为 running 并串行限频同步；终态为 succeeded/partial/failed。
    对接：BackgroundTasks；ChnenergyControlledClient；BidSourceHitRow。
    二次开发：禁止真实任意 URL、自动立项、持久化 Cookie/HTML/JSON；测试须注入 MockTransport 与零等待 sleep。
    """
    settings = get_settings()
    db = SessionLocal()
    client: ChnenergyControlledClient | None = None
    counters = {
        "plan_count": 0,
        "candidate_count": 0,
        "detail_page_count": 0,
        "resolved_count": 0,
        "needs_review_count": 0,
        "skipped_count": 0,
    }
    detail_cache: dict[str, dict[str, Any]] = {}
    hit_detail_limit = False
    try:
        run = db.scalars(
            select(BidSourceSyncRunRow).where(BidSourceSyncRunRow.id == run_id)
        ).first()
        if run is None:
            return
        if run.status not in ("queued", "running"):
            return

        now = _now()
        run.status = "running"
        run.updated_at = now
        db.commit()

        plans = _list_enabled_plans(
            db,
            run.workspace_id,
            limit=settings.max_opportunity_watch_plans_per_sync,
        )
        counters["plan_count"] = len(plans)

        client = ChnenergyControlledClient(
            transport=transport,
            sleep_fn=sleep_fn,
            min_interval_seconds=(
                settings.opportunity_watch_min_interval_seconds
                if min_interval_seconds is None
                else min_interval_seconds
            ),
            connect_timeout_seconds=settings.opportunity_watch_connect_timeout_seconds,
            read_timeout_seconds=settings.opportunity_watch_read_timeout_seconds,
            search_retry_count=(
                settings.opportunity_watch_search_retry_count
                if search_retry_count is None
                else search_retry_count
            ),
        )
        client.open()
        client.ensure_session()

        max_details = settings.max_opportunity_watch_detail_pages_per_sync
        max_candidates = settings.max_opportunity_watch_candidates_per_plan

        for plan in plans:
            try:
                records = client.search_candidates(plan.title)
            except ChnenergySyncStopError as stop:
                _finalize_run(
                    db,
                    run,
                    status="failed",
                    error_code=stop.error_code,
                    counters=counters,
                )
                return
            except ChnenergyNetworkError:
                # 单次检索网络失败（未达连续两次阈值）：跳过该计划
                counters["skipped_count"] += 1
                continue

            for record in records[:max_candidates]:
                counters["candidate_count"] += 1
                linkurl = record.get("linkurl") or ""
                title = record.get("title") or ""
                publish_text = record.get("infodate") or ""
                try:
                    fields = parse_jump_fields(linkurl)
                except ChnenergyClientError:
                    counters["skipped_count"] += 1
                    continue

                info_id = fields["infoid"]
                category_num = fields["categorynum"]
                date8 = fields["infodate"]

                if info_id not in detail_cache:
                    if counters["detail_page_count"] >= max_details:
                        hit_detail_limit = True
                        counters["skipped_count"] += 1
                        continue
                    try:
                        detail_url = build_notice_detail_url(
                            infoid=info_id,
                            categorynum=category_num,
                            infodate=date8,
                        )
                        html = client.fetch_detail_html(detail_url)
                        times = extract_notice_times(html)
                    except ChnenergySyncStopError as stop:
                        _finalize_run(
                            db,
                            run,
                            status="failed",
                            error_code=stop.error_code,
                            counters=counters,
                        )
                        return
                    except (ChnenergyNetworkError, ChnenergyClientError):
                        counters["skipped_count"] += 1
                        continue
                    finally:
                        html = ""  # 丢弃正文引用

                    detail_cache[info_id] = {
                        "title": title,
                        "publish_text": publish_text,
                        "category_num": category_num,
                        "deadline_at_local": times.get("deadline_at_local"),
                        "opening_at_local": times.get("opening_at_local"),
                        "extraction_status": times.get("extraction_status")
                        or "needs_review",
                    }
                    counters["detail_page_count"] += 1

                cached = detail_cache[info_id]
                status = cached["extraction_status"]
                _upsert_hit(
                    db,
                    workspace_id=run.workspace_id,
                    watch_plan_id=plan.id,
                    sync_run_id=run.id,
                    source_info_id=info_id,
                    category_num=cached["category_num"],
                    source_publish_text=cached["publish_text"] or publish_text,
                    title=cached["title"] or title,
                    deadline_at_local=cached["deadline_at_local"],
                    opening_at_local=cached["opening_at_local"],
                    extraction_status=status,
                )
                if status == "resolved":
                    counters["resolved_count"] += 1
                else:
                    counters["needs_review_count"] += 1

            db.commit()

        final_status = "partial" if hit_detail_limit else "succeeded"
        _finalize_run(
            db,
            run,
            status=final_status,
            error_code=None,
            counters=counters,
        )
    except ChnenergySyncStopError as stop:
        run = db.scalars(
            select(BidSourceSyncRunRow).where(BidSourceSyncRunRow.id == run_id)
        ).first()
        if run is not None and run.status in ("queued", "running"):
            _finalize_run(
                db,
                run,
                status="failed",
                error_code=stop.error_code,
                counters=counters,
            )
    except Exception:
        run = db.scalars(
            select(BidSourceSyncRunRow).where(BidSourceSyncRunRow.id == run_id)
        ).first()
        if run is not None and run.status in ("queued", "running"):
            _finalize_run(
                db,
                run,
                status="failed",
                error_code="source_unavailable",
                counters=counters,
            )
    finally:
        if client is not None:
            client.close()
        db.close()
